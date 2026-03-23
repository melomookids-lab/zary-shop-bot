import os
import re
import json
import html
import asyncio
import logging
import sqlite3
from pathlib import Path
from typing import Any, Optional
from datetime import datetime, timezone

from aiohttp import web
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message,
    CallbackQuery,
    BufferedInputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    WebAppInfo,
)

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "bot.db"
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
BASE_URL = os.getenv("BASE_URL", "").rstrip("/")
PORT = int(os.getenv("PORT", "8080"))
SHOP_BRAND = (os.getenv("SHOP_BRAND", "ZARY & CO") or "ZARY & CO").strip()
DEFAULT_LANGUAGE = (os.getenv("DEFAULT_LANGUAGE", "ru") or "ru").strip()
ADMIN_PANEL_TOKEN = os.getenv("ADMIN_PANEL_TOKEN", "").strip()
CHANNEL_LINK = os.getenv("CHANNEL_LINK", "").strip()
INSTAGRAM_LINK = os.getenv("INSTAGRAM_LINK", "").strip()
YOUTUBE_LINK = os.getenv("YOUTUBE_LINK", "").strip()
MANAGER_PHONE = os.getenv("MANAGER_PHONE", "+998771202255").strip()
MANAGER_TG = os.getenv("MANAGER_TG", "@manager").strip()
LOW_STOCK_THRESHOLD = int(os.getenv("LOW_STOCK_THRESHOLD", "3"))

ADMIN_IDS = {
    int(x.strip())
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x.strip() and x.strip().lstrip("-").isdigit()
}

SUPPORTED_LANGS = ("ru", "uz")
CATEGORY_SLUGS = ("new", "hits", "sale", "limited", "school", "casual")
ORDER_STATUSES = ("new", "confirmed", "paid", "sent", "delivered", "cancelled")
PAYMENT_STATUSES = ("pending", "paid", "failed", "cancelled", "refunded")
DELIVERY_METHODS = ("yandex_courier", "b2b_pochta", "yandex_pvz", "pickup")

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is not set")

if not BASE_URL:
    raise RuntimeError("BASE_URL is not set")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("telegram_shop_v2")

# ============================================================
# BOT / ROUTERS
# ============================================================

bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

user_router = Router()
cart_router = Router()
checkout_router = Router()
orders_router = Router()
reviews_router = Router()
admin_router = Router()
fallback_router = Router()
web_router = web.RouteTableDef()

# ============================================================
# TEXTS
# ============================================================

TEXTS = {
    "ru": {
        "menu_shop": "🛍 Магазин",
        "menu_cart": "🛒 Корзина",
        "menu_orders": "📦 Мои заказы",
        "menu_reviews": "⭐ Отзывы",
        "menu_leave_review": "✍️ Оставить отзыв",
        "menu_contacts": "📞 Контакты",
        "menu_lang": "🌐 Язык",
        "menu_admin": "🛠 Админ",
        "welcome": f"Добро пожаловать в <b>{SHOP_BRAND}</b>\n\nПремиальный shop bot внутри Telegram.",
        "main_menu_hint": "Выберите раздел ниже.",
        "choose_lang": "Выберите язык.",
        "lang_updated": "Язык обновлён.",
        "shop_opened": "Откройте магазин через кнопку ниже.",
        "cart_empty": "Корзина пуста.",
        "cart_added": "Товар добавлен в корзину.",
        "cart_updated": "Количество товара в корзине обновлено.",
        "cart_removed": "Позиция удалена из корзины.",
        "cart_cleared": "Корзина очищена.",
        "checkout_intro": "Начинаем оформление заказа.",
        "checkout_name": "Введите имя получателя.",
        "checkout_phone": "Введите телефон в формате +998901234567",
        "checkout_delivery": "Выберите доставку.",
        "checkout_city": "Введите город.",
        "checkout_address": "Введите адрес.",
        "checkout_payment": "Выберите способ оплаты.",
        "checkout_comment": "Введите комментарий или нажмите «Пропустить».",
        "checkout_summary": "🧾 <b>Проверка заказа</b>",
        "checkout_confirm_hint": "Напишите «Да» для подтверждения или «Нет» для отмены.",
        "checkout_done": "✅ Заказ создан.",
        "checkout_invalid_phone": "Некорректный номер телефона.",
        "cancel": "❌ Отмена",
        "skip": "Пропустить",
        "yes": "Да",
        "no": "Нет",
        "contacts_title": "📞 <b>Контакты</b>",
        "reviews_empty": "Отзывов пока нет.",
        "review_ask_rating": "Поставьте оценку от 1 до 5.",
        "review_ask_text": "Напишите текст отзыва.",
        "review_sent": "Спасибо, отзыв отправлен на модерацию.",
        "review_bad_rating": "Введите число от 1 до 5.",
        "not_admin": "Это доступно только админу.",
        "admin_title": "🛠 <b>Админ меню</b>",
        "admin_stats": "📊 Статистика",
        "admin_products": "📦 Товары",
        "admin_orders": "📋 Заказы",
        "admin_reviews": "⭐ Отзывы",
        "admin_stock": "📉 Остатки",
        "admin_report": "📁 Отчёт",
        "admin_customers": "👥 Клиенты",
        "admin_back": "⬅️ В меню",
        "delivery_yandex_courier": "Яндекс Курьер",
        "delivery_b2b_pochta": "B2B Почта",
        "delivery_yandex_pvz": "Яндекс ПВЗ",
        "delivery_pickup": "Самовывоз",
        "payment_click": "Click",
        "payment_payme": "Payme",
        "payment_cash": "Наличными",
        "status_new": "Новый",
        "status_confirmed": "Подтверждён",
        "status_paid": "Оплачен",
        "status_sent": "Отправлен",
        "status_delivered": "Доставлен",
        "status_cancelled": "Отменён",
        "payment_status_pending": "Ожидает оплаты",
        "payment_status_paid": "Оплачен",
        "payment_status_failed": "Ошибка",
        "payment_status_cancelled": "Отменён",
        "payment_status_refunded": "Возврат",
        "notify_order_new": "✅ Ваш заказ принят.",
        "notify_order_confirmed": "📦 Ваш заказ подтверждён.",
        "notify_order_paid": "💳 Ваш заказ оплачен.",
        "notify_order_sent": "🚚 Ваш заказ отправлен.",
        "notify_order_delivered": "🎉 Ваш заказ доставлен.",
        "stock_low": "⚠️ Мало на складе",
        "stock_out": "⛔ Нет в наличии",
    },
    "uz": {
        "menu_shop": "🛍 Do'kon",
        "menu_cart": "🛒 Savatcha",
        "menu_orders": "📦 Buyurtmalarim",
        "menu_reviews": "⭐ Sharhlar",
        "menu_leave_review": "✍️ Sharh qoldirish",
        "menu_contacts": "📞 Kontaktlar",
        "menu_lang": "🌐 Til",
        "menu_admin": "🛠 Admin",
        "welcome": f"<b>{SHOP_BRAND}</b> ga xush kelibsiz.\n\nPremium shop bot Telegram ichida.",
        "main_menu_hint": "Quyidagi bo'limni tanlang.",
        "choose_lang": "Tilni tanlang.",
        "lang_updated": "Til yangilandi.",
        "shop_opened": "Quyidagi tugma orqali do'konni oching.",
        "cart_empty": "Savatcha bo'sh.",
        "cart_added": "Mahsulot savatchaga qo'shildi.",
        "cart_updated": "Savatchadagi son yangilandi.",
        "cart_removed": "Pozitsiya o'chirildi.",
        "cart_cleared": "Savatcha tozalandi.",
        "checkout_intro": "Buyurtma rasmiylashtirish boshlandi.",
        "checkout_name": "Qabul qiluvchi ismini kiriting.",
        "checkout_phone": "Telefonni +998901234567 formatida kiriting.",
        "checkout_delivery": "Yetkazib berish usulini tanlang.",
        "checkout_city": "Shaharni kiriting.",
        "checkout_address": "Manzilni kiriting.",
        "checkout_payment": "To'lov usulini tanlang.",
        "checkout_comment": "Izoh kiriting yoki «O'tkazib yuborish» ni bosing.",
        "checkout_summary": "🧾 <b>Buyurtmani tekshirish</b>",
        "checkout_confirm_hint": "Tasdiqlash uchun «Ha», bekor qilish uchun «Yo'q» yuboring.",
        "checkout_done": "✅ Buyurtma yaratildi.",
        "checkout_invalid_phone": "Telefon noto'g'ri.",
        "cancel": "❌ Bekor qilish",
        "skip": "O'tkazib yuborish",
        "yes": "Ha",
        "no": "Yo'q",
        "contacts_title": "📞 <b>Kontaktlar</b>",
        "reviews_empty": "Hozircha sharhlar yo'q.",
        "review_ask_rating": "1 dan 5 gacha baho yuboring.",
        "review_ask_text": "Sharh matnini yozing.",
        "review_sent": "Rahmat, sharh moderatsiyaga yuborildi.",
        "review_bad_rating": "1 dan 5 gacha son kiriting.",
        "not_admin": "Bu faqat admin uchun.",
        "admin_title": "🛠 <b>Admin menyu</b>",
        "admin_stats": "📊 Statistika",
        "admin_products": "📦 Mahsulotlar",
        "admin_orders": "📋 Buyurtmalar",
        "admin_reviews": "⭐ Sharhlar",
        "admin_stock": "📉 Qoldiq",
        "admin_report": "📁 Hisobot",
        "admin_customers": "👥 Mijozlar",
        "admin_back": "⬅️ Menyuga",
        "delivery_yandex_courier": "Yandex Kuryer",
        "delivery_b2b_pochta": "B2B Pochta",
        "delivery_yandex_pvz": "Yandex PVZ",
        "delivery_pickup": "Olib ketish",
        "payment_click": "Click",
        "payment_payme": "Payme",
        "payment_cash": "Naqd",
        "status_new": "Yangi",
        "status_confirmed": "Tasdiqlangan",
        "status_paid": "To'langan",
        "status_sent": "Yuborilgan",
        "status_delivered": "Yetkazilgan",
        "status_cancelled": "Bekor qilingan",
        "payment_status_pending": "Kutilmoqda",
        "payment_status_paid": "To'langan",
        "payment_status_failed": "Xato",
        "payment_status_cancelled": "Bekor qilingan",
        "payment_status_refunded": "Qaytarilgan",
        "notify_order_new": "✅ Buyurtmangiz qabul qilindi.",
        "notify_order_confirmed": "📦 Buyurtmangiz tasdiqlandi.",
        "notify_order_paid": "💳 Buyurtmangiz to'landi.",
        "notify_order_sent": "🚚 Buyurtmangiz yuborildi.",
        "notify_order_delivered": "🎉 Buyurtmangiz yetkazildi.",
        "stock_low": "⚠️ Qoldiq kam",
        "stock_out": "⛔ Tugagan",
    },
}

# ============================================================
# DB
# ============================================================

def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def utc_now_iso() -> str:
    return utc_now().replace(microsecond=0).isoformat()


def init_db() -> None:
    conn = get_db()
    cur = conn.cursor()
    cur.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA foreign_keys=ON;

        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT NOT NULL DEFAULT '',
            full_name TEXT NOT NULL DEFAULT '',
            phone TEXT NOT NULL DEFAULT '',
            city TEXT NOT NULL DEFAULT '',
            lang TEXT NOT NULL DEFAULT 'ru',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title_ru TEXT NOT NULL,
            title_uz TEXT NOT NULL,
            description_ru TEXT NOT NULL DEFAULT '',
            description_uz TEXT NOT NULL DEFAULT '',
            photo_file_id TEXT NOT NULL DEFAULT '',
            category_slug TEXT NOT NULL DEFAULT 'casual',
            price INTEGER NOT NULL DEFAULT 0,
            old_price INTEGER NOT NULL DEFAULT 0,
            sizes TEXT NOT NULL DEFAULT '',
            stock_qty INTEGER NOT NULL DEFAULT 0,
            is_published INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 100,
            is_new INTEGER NOT NULL DEFAULT 0,
            is_hit INTEGER NOT NULL DEFAULT 0,
            is_limited INTEGER NOT NULL DEFAULT 0,
            discount_percent INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS carts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            size TEXT NOT NULL DEFAULT '',
            qty INTEGER NOT NULL DEFAULT 1,
            added_at TEXT NOT NULL,
            UNIQUE(user_id, product_id, size)
        );

        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT NOT NULL DEFAULT '',
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            customer_phone TEXT NOT NULL DEFAULT '',
            city TEXT NOT NULL DEFAULT '',
            items_json TEXT NOT NULL,
            total_qty INTEGER NOT NULL DEFAULT 0,
            total_amount INTEGER NOT NULL DEFAULT 0,
            delivery_method TEXT NOT NULL DEFAULT 'pickup',
            delivery_address TEXT NOT NULL DEFAULT '',
            payment_method TEXT NOT NULL DEFAULT 'cash',
            payment_status TEXT NOT NULL DEFAULT 'pending',
            status TEXT NOT NULL DEFAULT 'new',
            source TEXT NOT NULL DEFAULT 'telegram',
            comment TEXT NOT NULL DEFAULT '',
            is_new_customer INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS reviews (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            username TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            rating INTEGER NOT NULL DEFAULT 5,
            text TEXT NOT NULL DEFAULT '',
            is_published INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS favorites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(user_id, product_id)
        );

        CREATE INDEX IF NOT EXISTS idx_products_pub ON products(is_published, category_slug, sort_order);
        CREATE INDEX IF NOT EXISTS idx_orders_user ON orders(user_id);
        CREATE INDEX IF NOT EXISTS idx_orders_created ON orders(created_at);
        CREATE INDEX IF NOT EXISTS idx_reviews_pub ON reviews(is_published);
        CREATE INDEX IF NOT EXISTS idx_carts_user ON carts(user_id);
        """
    )
    conn.commit()
    conn.close()


def migrate_db() -> None:
    conn = get_db()
    cur = conn.cursor()

    def has_column(table: str, col: str) -> bool:
        rows = cur.execute(f"PRAGMA table_info({table})").fetchall()
        return any(r["name"] == col for r in rows)

    if not has_column("products", "is_new"):
        cur.execute("ALTER TABLE products ADD COLUMN is_new INTEGER NOT NULL DEFAULT 0")
    if not has_column("products", "is_hit"):
        cur.execute("ALTER TABLE products ADD COLUMN is_hit INTEGER NOT NULL DEFAULT 0")
    if not has_column("products", "is_limited"):
        cur.execute("ALTER TABLE products ADD COLUMN is_limited INTEGER NOT NULL DEFAULT 0")
    if not has_column("products", "discount_percent"):
        cur.execute("ALTER TABLE products ADD COLUMN discount_percent INTEGER NOT NULL DEFAULT 0")
    if not has_column("products", "sizes"):
        cur.execute("ALTER TABLE products ADD COLUMN sizes TEXT NOT NULL DEFAULT ''")
    if not has_column("orders", "order_number"):
        cur.execute("ALTER TABLE orders ADD COLUMN order_number TEXT NOT NULL DEFAULT ''")

    conn.commit()
    conn.close()


# ============================================================
# HELPERS
# ============================================================

def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def ensure_lang(lang: str) -> str:
    lang = (lang or "").strip()
    return lang if lang in SUPPORTED_LANGS else DEFAULT_LANGUAGE


def normalize_phone(phone: str) -> str:
    value = (phone or "").strip()
    value = value.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if value.startswith("998") and not value.startswith("+998"):
        value = "+" + value
    return value


def is_valid_phone(phone: str) -> bool:
    return bool(re.fullmatch(r"\+998\d{9}", normalize_phone(phone)))


def fmt_sum(value: Any) -> str:
    return f"{safe_int(value):,}".replace(",", " ") + " сум"


def split_sizes(value: str) -> list[str]:
    if not value:
        return []
    return [x.strip() for x in re.split(r"[,\n;/|]+", value) if x.strip()]


def badges_for_product(row: sqlite3.Row | dict[str, Any]) -> list[str]:
    result = []
    if safe_int(row["is_new"]):
        result.append("new")
    if safe_int(row["is_hit"]):
        result.append("hit")
    if safe_int(row["discount_percent"]) or safe_int(row["old_price"]) > safe_int(row["price"]):
        result.append("sale")
    if safe_int(row["is_limited"]):
        result.append("limited")
    stock = safe_int(row["stock_qty"])
    if 0 < stock <= LOW_STOCK_THRESHOLD:
        result.append("low_stock")
    return result


def get_user_lang(user_id: int) -> str:
    conn = get_db()
    row = conn.execute("SELECT lang FROM users WHERE user_id = ?", (user_id,)).fetchone()
    conn.close()
    return row["lang"] if row and row["lang"] in SUPPORTED_LANGS else DEFAULT_LANGUAGE


def t(lang_or_user: int | str, key: str) -> str:
    lang = get_user_lang(lang_or_user) if isinstance(lang_or_user, int) else ensure_lang(lang_or_user)
    return TEXTS.get(lang, TEXTS[DEFAULT_LANGUAGE]).get(key, key)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def escape(s: Any) -> str:
    return html.escape(str(s or ""))


def user_main_menu(user_id: int) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(text=t(user_id, "menu_shop"), web_app=WebAppInfo(url=f"{BASE_URL}/shop?lang={get_user_lang(user_id)}"))],
        [KeyboardButton(text=t(user_id, "menu_cart")), KeyboardButton(text=t(user_id, "menu_orders"))],
        [KeyboardButton(text=t(user_id, "menu_reviews")), KeyboardButton(text=t(user_id, "menu_leave_review"))],
        [KeyboardButton(text=t(user_id, "menu_contacts")), KeyboardButton(text=t(user_id, "menu_lang"))],
    ]
    if is_admin(user_id):
        rows.append([KeyboardButton(text=t(user_id, "menu_admin"))])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def admin_menu(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(user_id, "admin_stats")), KeyboardButton(text=t(user_id, "admin_stock"))],
            [KeyboardButton(text=t(user_id, "admin_orders")), KeyboardButton(text=t(user_id, "admin_products"))],
            [KeyboardButton(text=t(user_id, "admin_customers")), KeyboardButton(text=t(user_id, "admin_reviews"))],
            [KeyboardButton(text=t(user_id, "admin_report"))],
            [KeyboardButton(text=t(user_id, "admin_back"))],
        ],
        resize_keyboard=True,
    )


def language_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Русский", callback_data="lang:ru")],
            [InlineKeyboardButton(text="O'zbekcha", callback_data="lang:uz")],
        ]
    )


def cancel_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t(user_id, "cancel"))]],
        resize_keyboard=True,
    )


def yes_no_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t(user_id, "yes")), KeyboardButton(text=t(user_id, "no"))]],
        resize_keyboard=True,
    )


def delivery_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(user_id, "delivery_yandex_courier"))],
            [KeyboardButton(text=t(user_id, "delivery_b2b_pochta"))],
            [KeyboardButton(text=t(user_id, "delivery_yandex_pvz"))],
            [KeyboardButton(text=t(user_id, "delivery_pickup"))],
            [KeyboardButton(text=t(user_id, "cancel"))],
        ],
        resize_keyboard=True,
    )


def payment_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(user_id, "payment_click"))],
            [KeyboardButton(text=t(user_id, "payment_payme"))],
            [KeyboardButton(text=t(user_id, "payment_cash"))],
            [KeyboardButton(text=t(user_id, "cancel"))],
        ],
        resize_keyboard=True,
    )


def skip_keyboard(user_id: int) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=t(user_id, "skip"))], [KeyboardButton(text=t(user_id, "cancel"))]],
        resize_keyboard=True,
    )


def delivery_method_from_label(user_id: int, text: str) -> Optional[str]:
    mapping = {
        t(user_id, "delivery_yandex_courier"): "yandex_courier",
        t(user_id, "delivery_b2b_pochta"): "b2b_pochta",
        t(user_id, "delivery_yandex_pvz"): "yandex_pvz",
        t(user_id, "delivery_pickup"): "pickup",
    }
    return mapping.get((text or "").strip())


def payment_method_from_label(user_id: int, text: str) -> Optional[str]:
    mapping = {
        t(user_id, "payment_click"): "click",
        t(user_id, "payment_payme"): "payme",
        t(user_id, "payment_cash"): "cash",
    }
    return mapping.get((text or "").strip())


def delivery_label(lang_or_user: int | str, value: str) -> str:
    return t(lang_or_user, f"delivery_{value}")


def status_label(lang_or_user: int | str, value: str) -> str:
    return t(lang_or_user, f"status_{value}")


def payment_status_label(lang_or_user: int | str, value: str) -> str:
    return t(lang_or_user, f"payment_status_{value}")


def upsert_user(user_id: int, username: str | None, full_name: str | None) -> None:
    now = utc_now_iso()
    conn = get_db()
    row = conn.execute("SELECT lang FROM users WHERE user_id = ?", (user_id,)).fetchone()
    lang = row["lang"] if row and row["lang"] in SUPPORTED_LANGS else DEFAULT_LANGUAGE
    conn.execute(
        """
        INSERT INTO users (user_id, username, full_name, lang, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            username=excluded.username,
            full_name=excluded.full_name,
            updated_at=excluded.updated_at
        """,
        (user_id, username or "", full_name or "", lang, now, now),
    )
    conn.commit()
    conn.close()


def update_user_profile(user_id: int, phone: str = "", city: str = "") -> None:
    conn = get_db()
    conn.execute(
        "UPDATE users SET phone = CASE WHEN ? != '' THEN ? ELSE phone END, city = CASE WHEN ? != '' THEN ? ELSE city END, updated_at = ? WHERE user_id = ?",
        (phone, phone, city, city, utc_now_iso(), user_id),
    )
    conn.commit()
    conn.close()


def set_user_lang(user_id: int, lang: str) -> None:
    conn = get_db()
    conn.execute("UPDATE users SET lang = ?, updated_at = ? WHERE user_id = ?", (ensure_lang(lang), utc_now_iso(), user_id))
    conn.commit()
    conn.close()


def month_prefix(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


# ============================================================
# PRODUCT
# ============================================================

def get_products(published_only: bool = True) -> list[sqlite3.Row]:
    conn = get_db()
    if published_only:
        rows = conn.execute(
            "SELECT * FROM products WHERE is_published = 1 ORDER BY sort_order ASC, id DESC"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM products ORDER BY sort_order ASC, id DESC"
        ).fetchall()
    conn.close()
    return rows


def get_product(product_id: int) -> Optional[sqlite3.Row]:
    conn = get_db()
    row = conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    conn.close()
    return row


def product_title(row: sqlite3.Row | dict[str, Any], lang: str) -> str:
    return row["title_uz"] if ensure_lang(lang) == "uz" else row["title_ru"]


def product_description(row: sqlite3.Row | dict[str, Any], lang: str) -> str:
    return row["description_uz"] if ensure_lang(lang) == "uz" else row["description_ru"]


# ============================================================
# FAVORITES
# ============================================================

def toggle_favorite(user_id: int, product_id: int) -> bool:
    conn = get_db()
    row = conn.execute("SELECT id FROM favorites WHERE user_id = ? AND product_id = ?", (user_id, product_id)).fetchone()
    if row:
        conn.execute("DELETE FROM favorites WHERE id = ?", (row["id"],))
        conn.commit()
        conn.close()
        return False
    conn.execute("INSERT OR IGNORE INTO favorites (user_id, product_id, created_at) VALUES (?, ?, ?)", (user_id, product_id, utc_now_iso()))
    conn.commit()
    conn.close()
    return True


def get_user_favorites_set(user_id: int) -> set[int]:
    conn = get_db()
    rows = conn.execute("SELECT product_id FROM favorites WHERE user_id = ?", (user_id,)).fetchall()
    conn.close()
    return {safe_int(r["product_id"]) for r in rows}


# ============================================================
# CART
# ============================================================

def get_cart_rows(user_id: int) -> list[sqlite3.Row]:
    conn = get_db()
    rows = conn.execute(
        """
        SELECT
            c.id AS cart_id,
            c.user_id,
            c.product_id,
            c.size,
            c.qty,
            p.title_ru,
            p.title_uz,
            p.photo_file_id,
            p.price,
            p.stock_qty,
            p.old_price,
            p.is_new,
            p.is_hit,
            p.is_limited,
            p.discount_percent,
            p.is_published
        FROM carts c
        JOIN products p ON p.id = c.product_id
        WHERE c.user_id = ?
        ORDER BY c.id ASC
        """,
        (user_id,),
    ).fetchall()
    conn.close()
    return rows


def cart_totals(user_id: int) -> tuple[int, int]:
    rows = get_cart_rows(user_id)
    total_qty = sum(safe_int(r["qty"]) for r in rows)
    total_amount = sum(safe_int(r["qty"]) * safe_int(r["price"]) for r in rows)
    return total_qty, total_amount


def add_to_cart(user_id: int, product_id: int, qty: int = 1, size: str = "") -> tuple[bool, str]:
    product = get_product(product_id)
    if not product or safe_int(product["is_published"]) != 1:
        return False, "not_found"
    if safe_int(product["stock_qty"]) <= 0:
        return False, "stock_out"

    allowed_sizes = split_sizes(product["sizes"])
    if allowed_sizes and size and size not in allowed_sizes:
        return False, "bad_size"

    qty = max(1, safe_int(qty))
    conn = get_db()
    cur = conn.cursor()
    row = cur.execute(
        "SELECT * FROM carts WHERE user_id = ? AND product_id = ? AND size = ?",
        (user_id, product_id, size),
    ).fetchone()
    if row:
        new_qty = min(safe_int(product["stock_qty"]), safe_int(row["qty"]) + qty)
        cur.execute("UPDATE carts SET qty = ? WHERE id = ?", (new_qty, row["id"]))
        key = "cart_updated"
    else:
        cur.execute(
            "INSERT INTO carts (user_id, product_id, size, qty, added_at) VALUES (?, ?, ?, ?, ?)",
            (user_id, product_id, size, min(safe_int(product["stock_qty"]), qty), utc_now_iso()),
        )
        key = "cart_added"
    conn.commit()
    conn.close()
    return True, key


def remove_cart_item(user_id: int, cart_id: int) -> bool:
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM carts WHERE id = ? AND user_id = ?", (cart_id, user_id))
    ok = cur.rowcount > 0
    conn.commit()
    conn.close()
    return ok


def clear_cart(user_id: int) -> None:
    conn = get_db()
    conn.execute("DELETE FROM carts WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()


def cart_api(user_id: int, lang: str) -> dict[str, Any]:
    rows = get_cart_rows(user_id)
    items = []
    for row in rows:
        qty = safe_int(row["qty"])
        price = safe_int(row["price"])
        items.append({
            "cart_id": row["cart_id"],
            "product_id": row["product_id"],
            "product_name": row["title_uz"] if ensure_lang(lang) == "uz" else row["title_ru"],
            "qty": qty,
            "size": row["size"] or "",
            "price": price,
            "subtotal": qty * price,
        })
    total_qty, total_amount = cart_totals(user_id)
    return {"items": items, "total_qty": total_qty, "total_amount": total_amount}


# ============================================================
# ORDERS
# ============================================================

def user_has_previous_orders(user_id: int) -> bool:
    conn = get_db()
    row = conn.execute("SELECT COUNT(*) AS c FROM orders WHERE user_id = ?", (user_id,)).fetchone()
    conn.close()
    return safe_int(row["c"]) > 0


def build_order_number(order_id: int) -> str:
    return f"ORD-{datetime.now().strftime('%Y%m')}-{order_id:05d}"


def render_items_for_message(items_json: str) -> str:
    try:
        items = json.loads(items_json or "[]")
    except Exception:
        items = []
    if not items:
        return "—"
    lines = []
    for idx, item in enumerate(items, start=1):
        size = f" | {escape(item.get('size'))}" if item.get("size") else ""
        lines.append(
            f"{idx}. {escape(item.get('product_name'))}{size} — {safe_int(item.get('qty'))} × {fmt_sum(item.get('price'))}"
        )
    return "\n".join(lines)


def get_order(order_id: int) -> Optional[sqlite3.Row]:
    conn = get_db()
    row = conn.execute("SELECT * FROM orders WHERE id = ?", (order_id,)).fetchone()
    conn.close()
    return row


def get_orders_for_user(user_id: int, limit: int = 20) -> list[sqlite3.Row]:
    conn = get_db()
    rows = conn.execute("SELECT * FROM orders WHERE user_id = ? ORDER BY id DESC LIMIT ?", (user_id, limit)).fetchall()
    conn.close()
    return rows


def create_order(user_id: int, username: str, data: dict[str, Any], source: str = "telegram") -> int:
    cart_rows = get_cart_rows(user_id)
    if not cart_rows:
        raise ValueError("Cart is empty")

    conn = get_db()
    cur = conn.cursor()
    total_qty = 0
    total_amount = 0
    items: list[dict[str, Any]] = []

    for row in cart_rows:
        product = cur.execute("SELECT * FROM products WHERE id = ?", (row["product_id"],)).fetchone()
        if not product:
            continue

        stock = safe_int(product["stock_qty"])
        qty = min(stock, safe_int(row["qty"]))
        if qty <= 0:
            continue

        price = safe_int(product["price"])
        subtotal = qty * price
        total_qty += qty
        total_amount += subtotal

        items.append({
            "product_id": safe_int(product["id"]),
            "product_name": product["title_ru"],
            "size": row["size"] or "",
            "qty": qty,
            "price": price,
            "subtotal": subtotal,
        })

        new_stock = max(0, stock - qty)
        cur.execute(
            "UPDATE products SET stock_qty = ?, updated_at = ? WHERE id = ?",
            (new_stock, utc_now_iso(), product["id"]),
        )

    if not items:
        conn.close()
        raise ValueError("No available items in cart")

    payment_method = data.get("payment_method") or "cash"
    payment_status = "pending" if payment_method in {"click", "payme"} else "paid"
    is_new_customer = 0 if user_has_previous_orders(user_id) else 1
    now = utc_now_iso()

    cur.execute(
        """
        INSERT INTO orders (
            order_number, user_id, username, customer_name, customer_phone, city,
            items_json, total_qty, total_amount, delivery_method, delivery_address,
            payment_method, payment_status, status, source, comment,
            is_new_customer, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "",
            user_id,
            username or "",
            data.get("customer_name") or "",
            data.get("customer_phone") or "",
            data.get("city") or "",
            json.dumps(items, ensure_ascii=False),
            total_qty,
            total_amount,
            data.get("delivery_method") or "pickup",
            data.get("delivery_address") or "",
            payment_method,
            payment_status,
            "new",
            source,
            data.get("comment") or "",
            is_new_customer,
            now,
            now,
        ),
    )
    order_id = cur.lastrowid
    order_number = build_order_number(order_id)
    cur.execute("UPDATE orders SET order_number = ? WHERE id = ?", (order_number, order_id))
    cur.execute("DELETE FROM carts WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()
    return order_id


# ============================================================
# REVIEWS
# ============================================================

def add_review(user_id: int, username: str, customer_name: str, rating: int, text: str) -> int:
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO reviews (user_id, username, customer_name, rating, text, is_published, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, 0, ?, ?)
        """,
        (user_id, username or "", customer_name or "", rating, text, utc_now_iso(), utc_now_iso()),
    )
    review_id = cur.lastrowid
    conn.commit()
    conn.close()
    return review_id


# ============================================================
# REPORTS
# ============================================================

def monthly_report(year: int, month: int) -> tuple[Path, str]:
    prefix = month_prefix(year, month)
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM orders WHERE substr(created_at, 1, 7) = ? ORDER BY id ASC",
        (prefix,),
    ).fetchall()

    total_turnover = sum(safe_int(r["total_amount"]) for r in rows)
    total_orders = len(rows)
    new_customers = sum(safe_int(r["is_new_customer"]) for r in rows)

    product_counter: dict[str, int] = {}
    size_counter: dict[str, int] = {}
    delivery_counter: dict[str, int] = {}

    for row in rows:
        delivery_counter[row["delivery_method"]] = delivery_counter.get(row["delivery_method"], 0) + 1
        try:
            items = json.loads(row["items_json"] or "[]")
        except Exception:
            items = []
        for item in items:
            name = str(item.get("product_name") or "—")
            size = str(item.get("size") or "—")
            qty = safe_int(item.get("qty"))
            product_counter[name] = product_counter.get(name, 0) + qty
            size_counter[size] = size_counter.get(size, 0) + qty

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = [
        "Номер заказа", "Дата", "Имя клиента", "Телефон", "Город", "Товары",
        "Размеры", "Количество", "Сумма", "Доставка", "Способ оплаты",
        "Статус оплаты", "Статус заказа", "Источник заказа"
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="D9C28A")
    font = Font(bold=True)

    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        try:
            items = json.loads(row["items_json"] or "[]")
        except Exception:
            items = []
        products_text = ", ".join(str(i.get("product_name") or "—") for i in items)
        sizes_text = ", ".join(str(i.get("size") or "—") for i in items)
        qty = sum(safe_int(i.get("qty")) for i in items)

        ws.append([
            row["order_number"] or row["id"],
            row["created_at"],
            row["customer_name"],
            row["customer_phone"],
            row["city"],
            products_text,
            sizes_text,
            qty,
            safe_int(row["total_amount"]),
            row["delivery_method"],
            row["payment_method"],
            row["payment_status"],
            row["status"],
            row["source"],
        ])

    summary = wb.create_sheet("Summary")
    summary.append(["Метрика", "Значение"])
    summary["A1"].font = font
    summary["B1"].font = font
    summary["A1"].fill = fill
    summary["B1"].fill = fill

    summary.append(["Общий оборот за месяц", total_turnover])
    summary.append(["Количество заказов", total_orders])
    summary.append(["Новых клиентов", new_customers])
    summary.append([])
    summary.append(["Самые продаваемые товары", "Количество"])
    for name, qty in sorted(product_counter.items(), key=lambda x: x[1], reverse=True)[:10]:
        summary.append([name, qty])
    summary.append([])
    summary.append(["Самые популярные размеры", "Количество"])
    for size, qty in sorted(size_counter.items(), key=lambda x: x[1], reverse=True)[:10]:
        summary.append([size, qty])
    summary.append([])
    summary.append(["Сколько доставок каждым способом", "Количество"])
    for name, qty in sorted(delivery_counter.items(), key=lambda x: x[1], reverse=True):
        summary.append([name, qty])

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 22
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 18

    file_path = REPORTS_DIR / f"orders_report_{year}_{month:02d}.xlsx"
    wb.save(file_path)
    conn.close()

    caption = (
        f"📁 Отчёт за {month:02d}.{year}\n"
        f"Оборот: {fmt_sum(total_turnover)}\n"
        f"Заказов: {total_orders}\n"
        f"Новых клиентов: {new_customers}"
    )
    return file_path, caption


# ============================================================
# ADMIN / NOTIFICATIONS
# ============================================================

def get_admin_stats() -> dict[str, int]:
    conn = get_db()
    cur = conn.cursor()

    def scalar(query: str, params: tuple = ()) -> int:
        row = cur.execute(query, params).fetchone()
        return safe_int(row[0]) if row else 0

    stats = {
        "orders": scalar("SELECT COUNT(*) FROM orders"),
        "users": scalar("SELECT COUNT(*) FROM users"),
        "products": scalar("SELECT COUNT(*) FROM products"),
        "reviews": scalar("SELECT COUNT(*) FROM reviews WHERE is_published = 1"),
        "new_orders": scalar("SELECT COUNT(*) FROM orders WHERE status = 'new'"),
        "low_stock": scalar("SELECT COUNT(*) FROM products WHERE stock_qty <= ?", (LOW_STOCK_THRESHOLD,)),
    }
    conn.close()
    return stats


def low_stock_products() -> list[sqlite3.Row]:
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM products WHERE stock_qty <= ? ORDER BY stock_qty ASC, id DESC",
        (LOW_STOCK_THRESHOLD,),
    ).fetchall()
    conn.close()
    return rows


async def notify_admins_low_stock() -> None:
    rows = low_stock_products()
    if not rows or not ADMIN_IDS:
        return
    lines = ["📉 <b>Контроль склада</b>"]
    for row in rows[:20]:
        title = escape(row["title_ru"])
        stock = safe_int(row["stock_qty"])
        label = "⛔ Закончился" if stock <= 0 else f"⚠️ Осталось {stock} шт"
        lines.append(f"• #{row['id']} {title} — {label}")
    text = "\n".join(lines)
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text)
        except Exception:
            logger.exception("Failed to notify low stock to admin %s", admin_id)


async def notify_admins_new_order(order_id: int) -> None:
    order = get_order(order_id)
    if not order:
        return
    text = (
        f"📦 <b>Новый заказ {escape(order['order_number'] or order['id'])}</b>\n\n"
        f"<b>Клиент:</b> {escape(order['customer_name'])}\n"
        f"<b>Телефон:</b> {escape(order['customer_phone'])}\n"
        f"<b>Город:</b> {escape(order['city'])}\n"
        f"<b>Доставка:</b> {delivery_label('ru', order['delivery_method'])}\n"
        f"<b>Оплата:</b> {escape(order['payment_method'])} / {payment_status_label('ru', order['payment_status'])}\n"
        f"<b>Сумма:</b> {fmt_sum(order['total_amount'])}\n\n"
        f"<b>Товары:</b>\n{render_items_for_message(order['items_json'])}"
    )
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text)
        except Exception:
            logger.exception("Failed to notify admin %s about order %s", admin_id, order_id)


async def notify_user_order_status(order_id: int, status: str) -> None:
    order = get_order(order_id)
    if not order:
        return
    if status not in ORDER_STATUSES:
        return
    lang = get_user_lang(order["user_id"])
    mapping = {
        "new": "notify_order_new",
        "confirmed": "notify_order_confirmed",
        "paid": "notify_order_paid",
        "sent": "notify_order_sent",
        "delivered": "notify_order_delivered",
    }
    key = mapping.get(status)
    if not key:
        return
    text = (
        f"{t(lang, key)}\n\n"
        f"<b>Заказ:</b> {escape(order['order_number'] or order['id'])}\n"
        f"<b>Статус:</b> {status_label(lang, status)}\n"
        f"<b>Сумма:</b> {fmt_sum(order['total_amount'])}"
    )
    try:
        await bot.send_message(order["user_id"], text)
    except Exception:
        logger.exception("Failed to notify user %s about order %s status", order["user_id"], order_id)


# ============================================================
# FSM
# ============================================================

class ReviewState(StatesGroup):
    rating = State()
    text = State()


class CheckoutState(StatesGroup):
    customer_name = State()
    customer_phone = State()
    delivery_method = State()
    city = State()
    delivery_address = State()
    payment_method = State()
    comment = State()
    confirm = State()


# ============================================================
# FILE URL
# ============================================================

async def file_url(file_id: str) -> str:
    if not file_id:
        return ""
    try:
        file_info = await bot.get_file(file_id)
        return f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_info.file_path}"
    except Exception:
        return ""


def product_to_web_dict(row: sqlite3.Row, lang: str, photo_url: str, favorite_ids: set[int]) -> dict[str, Any]:
    stock = safe_int(row["stock_qty"])
    old_price = safe_int(row["old_price"])
    discount_percent = safe_int(row["discount_percent"])
    if discount_percent <= 0 and old_price > safe_int(row["price"]) and old_price > 0:
        discount_percent = max(1, int(round((old_price - safe_int(row["price"])) * 100 / old_price)))
    return {
        "id": safe_int(row["id"]),
        "title": product_title(row, lang),
        "description": product_description(row, lang),
        "photo_url": photo_url,
        "category_slug": row["category_slug"],
        "price": safe_int(row["price"]),
        "old_price": old_price,
        "discount_percent": discount_percent,
        "sizes_list": split_sizes(row["sizes"]),
        "stock_qty": stock,
        "can_buy": stock > 0 and safe_int(row["is_published"]) == 1,
        "badges": badges_for_product(row),
        "favorite": safe_int(row["id"]) in favorite_ids,
    }


# ============================================================
# WEBAPP HTML
# ============================================================

def build_shop_html() -> str:
    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, viewport-fit=cover">
<title>{escape(SHOP_BRAND)}</title>
<script src="https://telegram.org/js/telegram-web-app.js"></script>
<style>
:root {{
    --bg1:#fffaf4;
    --bg2:#fdf1f7;
    --bg3:#fcf7ee;
    --cream:#fffaf2;
    --ivory:#fffdf8;
    --champagne:#f6ebd7;
    --gold:#c7a865;
    --gold-deep:#9e7840;
    --text:#221912;
    --muted:#7c6c60;
    --pink:#b51463;
    --violet:#6f3dd9;
    --rose:#8f0f56;
    --glass:rgba(255,255,255,.62);
    --line:rgba(205,176,112,.24);
    --shadow:0 18px 50px rgba(96,70,32,.08);
    --radius:26px;
}}
* {{ box-sizing:border-box; }}
html,body {{ margin:0; padding:0; font-family:Inter,Arial,sans-serif; color:var(--text); background:
    radial-gradient(circle at top right, rgba(181,20,99,.08), transparent 28%),
    radial-gradient(circle at top left, rgba(111,61,217,.09), transparent 22%),
    linear-gradient(135deg,var(--bg1),var(--bg2),var(--bg3)); min-height:100%; }}
body {{ overflow-x:hidden; }}
a {{ color:inherit; text-decoration:none; }}
button,input,select {{ font:inherit; }}
.wrap {{ max-width:1240px; margin:0 auto; padding:18px 14px 32px; position:relative; z-index:2; }}
.hero {{
    position:relative;
    overflow:hidden;
    border-radius:0 0 34px 34px;
    padding:30px 22px 24px;
    background:
        linear-gradient(180deg, rgba(255,255,255,.74), rgba(255,255,255,.55)),
        radial-gradient(circle at 20% 20%, rgba(255,214,229,.45), transparent 34%),
        radial-gradient(circle at 80% 0%, rgba(218,203,255,.38), transparent 28%),
        linear-gradient(120deg,#fffaf4,#fff5f9,#fffaf0);
    backdrop-filter: blur(16px);
    border:1px solid rgba(255,255,255,.78);
    box-shadow:0 24px 64px rgba(86,60,26,.08);
}}
.brand {{
    font-size:40px;
    font-weight:900;
    text-transform:uppercase;
    letter-spacing:.15em;
    line-height:1;
    background:linear-gradient(180deg,#f8efcb,#c7a865,#9e7840);
    -webkit-background-clip:text;
    -webkit-text-fill-color:transparent;
}}
.subtitle {{
    color:var(--muted);
    margin-top:10px;
    max-width:560px;
    font-size:14px;
    line-height:1.5;
}}
.floating-shine {{
    position:absolute;
    inset:auto -40px -40px auto;
    width:220px;
    height:220px;
    border-radius:50%;
    background:radial-gradient(circle, rgba(255,235,190,.38), rgba(255,255,255,0));
    filter:blur(12px);
    pointer-events:none;
}}
.petals {{ position:fixed; inset:0; pointer-events:none; overflow:hidden; z-index:1; }}
.petal {{
    position:absolute;
    top:-80px;
    width:22px;
    height:28px;
    opacity:.9;
    filter:drop-shadow(0 6px 10px rgba(139,30,88,.18));
    border-radius:70% 0 70% 0;
    transform:rotate(25deg);
    animation-name:fall;
    animation-timing-function:linear;
    animation-iteration-count:infinite;
}}
.petal.v1 {{ background:linear-gradient(180deg,#8d39ff,#5a20b0); }}
.petal.v2 {{ background:linear-gradient(180deg,#d61f74,#861044); }}
.petal.v3 {{ background:linear-gradient(180deg,#ef9fc8,#b94382); }}
.petal.v4 {{ background:linear-gradient(180deg,#8e59ff,#ca4b89); }}
@keyframes fall {{
    0% {{ transform:translate3d(0,-80px,0) rotate(0deg) scale(.8); opacity:0; }}
    8% {{ opacity:.95; }}
    50% {{ transform:translate3d(40px,55vh,0) rotate(160deg) scale(1); }}
    100% {{ transform:translate3d(-30px,110vh,0) rotate(320deg) scale(.9); opacity:.15; }}
}}
.layout {{ display:grid; grid-template-columns:1.6fr .95fr; gap:16px; margin-top:16px; }}
.panel {{
    background:var(--glass);
    border:1px solid rgba(255,255,255,.72);
    backdrop-filter:blur(16px);
    border-radius:var(--radius);
    box-shadow:var(--shadow);
    overflow:hidden;
}}
.head {{ padding:20px 18px 4px; font-size:25px; font-weight:900; }}
.subhead {{ padding:0 18px 14px; color:var(--muted); font-size:14px; }}
.tools {{ padding:0 16px 14px; display:grid; grid-template-columns:1.2fr .85fr .85fr .85fr; gap:10px; }}
.search, .select, .input {{
    width:100%;
    border:none;
    outline:none;
    border-radius:16px;
    background:rgba(255,255,255,.88);
    padding:13px 14px;
    border:1px solid var(--line);
    color:var(--text);
}}
.filter-row {{ display:flex; gap:8px; flex-wrap:wrap; padding:0 16px 14px; }}
.chip {{
    border:none;
    cursor:pointer;
    padding:10px 14px;
    border-radius:999px;
    background:rgba(255,255,255,.88);
    border:1px solid var(--line);
    font-weight:800;
    color:#54483f;
}}
.chip.active {{
    color:#fff;
    background:linear-gradient(180deg,var(--gold),var(--gold-deep));
}}
.grid {{
    display:grid;
    grid-template-columns:repeat(2,minmax(0,1fr));
    gap:14px;
    padding:0 14px 16px;
}}
.card {{
    background:rgba(255,255,255,.8);
    border:1px solid rgba(255,255,255,.9);
    border-radius:24px;
    box-shadow:0 14px 30px rgba(96,70,32,.06);
    overflow:hidden;
}}
.card-inner {{ padding:14px; display:flex; flex-direction:column; gap:12px; }}
.photo-wrap {{ position:relative; }}
.photo {{
    aspect-ratio:1/1.12;
    border-radius:18px;
    overflow:hidden;
    background:
        radial-gradient(circle at top right, rgba(255,216,235,.36), transparent 32%),
        radial-gradient(circle at bottom left, rgba(219,204,255,.36), transparent 28%),
        linear-gradient(135deg,#fffbf7,#f5ecdf);
    display:flex; align-items:center; justify-content:center;
}}
.photo img {{ width:100%; height:100%; object-fit:cover; display:block; }}
.badges {{
    position:absolute;
    top:10px;
    left:10px;
    display:flex;
    gap:6px;
    flex-wrap:wrap;
    max-width:70%;
}}
.badge {{
    font-size:11px;
    font-weight:900;
    color:#fff;
    padding:6px 10px;
    border-radius:999px;
    box-shadow:0 8px 18px rgba(43,26,31,.16);
}}
.badge.new {{ background:linear-gradient(180deg,#7d49ff,#5920bf); }}
.badge.hit {{ background:linear-gradient(180deg,#d91d73,#8c1048); }}
.badge.sale {{ background:linear-gradient(180deg,#d1af68,#9c7330); }}
.badge.limited {{ background:linear-gradient(180deg,#19130f,#3b2a1c); }}
.badge.low_stock {{ background:linear-gradient(180deg,#ba0d5d,#7f103f); }}
.fav {{
    position:absolute;
    right:10px;
    top:10px;
    width:40px;
    height:40px;
    border:none;
    cursor:pointer;
    border-radius:50%;
    background:rgba(255,255,255,.9);
    border:1px solid rgba(255,255,255,.9);
    box-shadow:0 8px 18px rgba(96,70,32,.12);
    font-size:20px;
}}
.fav.active {{ background:linear-gradient(180deg,#ffd2e6,#fff); }}
.title {{ font-size:18px; font-weight:900; line-height:1.35; }}
.desc {{ font-size:13px; color:var(--muted); line-height:1.5; min-height:38px; }}
.price-row {{ display:flex; align-items:center; gap:8px; flex-wrap:wrap; }}
.price {{ font-size:22px; font-weight:900; }}
.old-price {{ text-decoration:line-through; color:#9a897e; font-size:13px; }}
.discount-pill {{
    background:rgba(214,31,116,.11);
    color:#9a0f52;
    font-weight:900;
    font-size:12px;
    border-radius:999px;
    padding:5px 10px;
}}
.stock {{
    font-size:13px;
    color:#6d5c51;
}}
.stock.low {{ color:#9c0d51; font-weight:800; }}
.stock.out {{ color:#6a5c52; font-weight:800; }}
.sizes {{ display:flex; gap:7px; flex-wrap:wrap; }}
.size-btn {{
    border:none;
    cursor:pointer;
    padding:8px 12px;
    border-radius:999px;
    background:#fff;
    border:1px solid var(--line);
    font-weight:800;
}}
.size-btn.active {{
    background:linear-gradient(180deg,var(--gold),var(--gold-deep));
    color:#fff;
}}
.actions {{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:10px;
}}
.btn-dark,.btn-gold,.btn-ghost {{
    border:none;
    cursor:pointer;
    border-radius:16px;
    padding:13px 14px;
    font-weight:900;
}}
.btn-dark {{ background:#19130f; color:#fff; }}
.btn-gold {{ background:linear-gradient(180deg,var(--gold),var(--gold-deep)); color:#fff; }}
.btn-ghost {{ background:rgba(255,255,255,.88); border:1px solid var(--line); color:var(--text); }}
.btn-dark:disabled,.btn-gold:disabled,.btn-ghost:disabled {{
    cursor:not-allowed;
    opacity:.45;
}}
.right-top {{
    padding:18px 16px 10px;
    display:flex; justify-content:space-between; align-items:center; gap:10px;
}}
.right-title {{ font-size:24px; font-weight:900; }}
.cart-list {{ display:flex; flex-direction:column; gap:10px; padding:0 16px 16px; }}
.cart-item {{
    padding:12px;
    border-radius:18px;
    background:rgba(255,255,255,.9);
    border:1px solid rgba(205,176,112,.18);
}}
.summary {{ padding:0 16px 18px; }}
.summary-row {{ display:flex; align-items:center; justify-content:space-between; margin-bottom:9px; }}
.reviews-box {{ padding:0 16px 18px; display:flex; flex-direction:column; gap:10px; }}
.review {{
    background:rgba(255,255,255,.85);
    border:1px solid rgba(205,176,112,.15);
    border-radius:18px;
    padding:12px;
}}
.review .name {{ font-weight:900; margin-bottom:6px; }}
.review .text {{ color:#65584e; line-height:1.5; font-size:14px; }}
.notice {{
    position:fixed;
    left:50%;
    bottom:18px;
    transform:translateX(-50%);
    background:#19130f;
    color:#fff;
    padding:12px 16px;
    border-radius:999px;
    opacity:0;
    transition:.2s ease;
    z-index:20;
    box-shadow:0 14px 28px rgba(0,0,0,.14);
}}
.notice.show {{ opacity:1; }}
.empty {{
    padding:18px;
    color:#6f6156;
}}
.mini-note {{ font-size:12px; color:#8a7a6f; margin-top:4px; }}
@media(max-width:1080px) {{
    .layout {{ grid-template-columns:1fr; }}
}}
@media(max-width:760px) {{
    .tools {{ grid-template-columns:1fr 1fr; }}
    .grid {{ grid-template-columns:1fr; }}
}}
@media(max-width:520px) {{
    .brand {{ font-size:30px; }}
    .tools {{ grid-template-columns:1fr; }}
    .actions {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class="petals" id="petals"></div>

<div class="wrap">
    <div class="hero">
        <div class="brand">{escape(SHOP_BRAND)}</div>
        <div class="subtitle">Premium selection with luxury spring mood, elegant glass layers and fast checkout inside Telegram.</div>
        <div class="floating-shine"></div>
    </div>

    <div class="layout">
        <div class="panel">
            <div class="head">Каталог</div>
            <div class="subhead">Поиск, категория, размер, цена, избранное и быстрый checkout.</div>

            <div class="tools">
                <input id="searchInput" class="search" placeholder="Поиск по названию">
                <select id="sizeFilter" class="select">
                    <option value="">Все размеры</option>
                </select>
                <input id="priceMin" class="input" placeholder="Цена от" inputmode="numeric">
                <input id="priceMax" class="input" placeholder="Цена до" inputmode="numeric">
            </div>

            <div class="filter-row" id="categoryFilters"></div>
            <div class="grid" id="productGrid"></div>
        </div>

        <div class="panel">
            <div class="right-top">
                <div class="right-title">Корзина</div>
                <button class="btn-ghost" id="favOnlyBtn">❤️ Избранное</button>
            </div>
            <div class="subhead">Оформление продолжается в боте.</div>
            <div class="cart-list" id="cartList"></div>
            <div class="summary">
                <div class="summary-row"><span>Всего товаров</span><b id="sumQty">0</b></div>
                <div class="summary-row"><span>Сумма</span><b id="sumAmount">0 сум</b></div>
                <button class="btn-dark" style="width:100%;margin-top:8px" id="checkoutBtn">Оформить заказ</button>
                <button class="btn-ghost" style="width:100%;margin-top:8px" id="clearBtn">Очистить корзину</button>
            </div>

            <div class="head" style="padding-top:8px">Отзывы</div>
            <div class="reviews-box" id="reviewsBox"></div>
        </div>
    </div>
</div>

<div class="notice" id="notice"></div>

<script>
const tg = window.Telegram?.WebApp || null;
if (tg) {{
    tg.ready();
    tg.expand();
}}

const qs = new URLSearchParams(location.search);
const lang = qs.get("lang") || "ru";
const userId = tg?.initDataUnsafe?.user?.id || 0;

const state = {{
    products: [],
    filtered: [],
    cart: [],
    reviews: [],
    favoritesOnly: false,
    category: "all",
    search: "",
    size: "",
    priceMin: "",
    priceMax: ""
}};

const I18N = {{
    ru: {{
        all: "Все",
        new: "Новинки",
        hits: "Хиты",
        sale: "Скидки",
        limited: "Limited",
        school: "School",
        casual: "Casual",
        add: "В корзину",
        buy: "Купить сейчас",
        out: "Нет в наличии",
        left: "Осталось",
        empty: "Ничего не найдено",
        cartEmpty: "Корзина пуста",
        reviewsEmpty: "Пока нет опубликованных отзывов",
        added: "Добавлено в корзину",
        removed: "Удалено из корзины",
        cleared: "Корзина очищена",
        favoriteAdded: "Добавлено в избранное",
        favoriteRemoved: "Удалено из избранного",
        checkout: "Переходим к оформлению"
    }},
    uz: {{
        all: "Barchasi",
        new: "Yangilar",
        hits: "Hitlar",
        sale: "Chegirma",
        limited: "Limited",
        school: "School",
        casual: "Casual",
        add: "Savatchaga",
        buy: "Hozir olish",
        out: "Sotuvda yo'q",
        left: "Qoldi",
        empty: "Hech narsa topilmadi",
        cartEmpty: "Savatcha bo'sh",
        reviewsEmpty: "Hozircha sharhlar yo'q",
        added: "Savatchaga qo'shildi",
        removed: "Savatchadan o'chirildi",
        cleared: "Savatcha tozalandi",
        favoriteAdded: "Saralanganlarga qo'shildi",
        favoriteRemoved: "Saralanganlardan o'chirildi",
        checkout: "Rasmiylashtirishga o'tamiz"
    }}
}};

const L = I18N[lang] || I18N.ru;

const CATEGORIES = [
    {{ key: "all", title: L.all }},
    {{ key: "new", title: L.new }},
    {{ key: "hits", title: L.hits }},
    {{ key: "sale", title: L.sale }},
    {{ key: "limited", title: L.limited }},
    {{ key: "school", title: L.school }},
    {{ key: "casual", title: L.casual }}
];

const BADGE_LABELS = {{
    ru: {{
        new: "Новинка",
        hit: "Хит",
        sale: "Скидка",
        limited: "Limited",
        low_stock: "Скоро закончится"
    }},
    uz: {{
        new: "Yangi",
        hit: "Hit",
        sale: "Chegirma",
        limited: "Limited",
        low_stock: "Tez tugaydi"
    }}
}};

function esc(v) {{
    return String(v ?? "").replace(/[&<>"']/g, s => ({{"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#039;"}}[s]));
}}

function fmtSum(v) {{
    return Number(v || 0).toLocaleString("ru-RU") + " сум";
}}

function notice(text) {{
    const box = document.getElementById("notice");
    box.textContent = text;
    box.classList.add("show");
    setTimeout(() => box.classList.remove("show"), 1900);
}}

function sendData(payload) {{
    if (!tg) {{
        notice("Open inside Telegram");
        return;
    }}
    tg.sendData(JSON.stringify(payload));
}}

function createPetals() {{
    const root = document.getElementById("petals");
    root.innerHTML = "";
    const classes = ["v1", "v2", "v3", "v4"];
    for (let i = 0; i < 28; i++) {{
        const el = document.createElement("div");
        el.className = "petal " + classes[i % classes.length];
        el.style.left = Math.random() * 100 + "vw";
        el.style.animationDuration = (8 + Math.random() * 10) + "s";
        el.style.animationDelay = (-Math.random() * 15) + "s";
        el.style.opacity = String(0.45 + Math.random() * 0.5);
        el.style.transform = "rotate(" + Math.floor(Math.random() * 360) + "deg)";
        root.appendChild(el);
    }}
}}

async function api(path, opts = undefined) {{
    const res = await fetch(path, opts);
    return await res.json();
}}

async function loadProducts() {{
    const data = await api(`/api/shop/products?lang=${{encodeURIComponent(lang)}}&user_id=${{encodeURIComponent(userId)}}`);
    state.products = Array.isArray(data) ? data : [];
    buildSizeFilter();
    applyFilters();
}}

async function loadCart() {{
    if (!userId) {{
        state.cart = [];
        renderCart();
        return;
    }}
    const data = await api(`/api/shop/cart?user_id=${{encodeURIComponent(userId)}}&lang=${{encodeURIComponent(lang)}}`);
    state.cart = data.items || [];
    renderCart();
}}

async function loadReviews() {{
    const data = await api(`/api/shop/reviews`);
    state.reviews = Array.isArray(data) ? data : [];
    renderReviews();
}}

function buildCategoryFilters() {{
    const box = document.getElementById("categoryFilters");
    box.innerHTML = "";
    CATEGORIES.forEach(cat => {{
        const btn = document.createElement("button");
        btn.className = "chip" + (state.category === cat.key ? " active" : "");
        btn.textContent = cat.title;
        btn.onclick = () => {{
            state.category = cat.key;
            buildCategoryFilters();
            applyFilters();
        }};
        box.appendChild(btn);
    }});
}}

function buildSizeFilter() {{
    const allSizes = new Set();
    state.products.forEach(p => (p.sizes_list || []).forEach(s => allSizes.add(s)));
    const sel = document.getElementById("sizeFilter");
    const current = state.size;
    sel.innerHTML = `<option value="">${{lang === "uz" ? "Barcha o'lchamlar" : "Все размеры"}}</option>`;
    Array.from(allSizes).sort().forEach(size => {{
        const opt = document.createElement("option");
        opt.value = size;
        opt.textContent = size;
        if (size === current) opt.selected = true;
        sel.appendChild(opt);
    }});
}}

function applyFilters() {{
    const q = state.search.trim().toLowerCase();
    const size = state.size;
    const min = state.priceMin ? Number(state.priceMin) : 0;
    const max = state.priceMax ? Number(state.priceMax) : 0;

    state.filtered = state.products.filter(p => {{
        const textMatch = !q || String(p.title || "").toLowerCase().includes(q);
        let categoryMatch = state.category === "all";
        if (!categoryMatch) {{
            if (state.category === "new") categoryMatch = !!(p.badges || []).includes("new");
            else if (state.category === "hits") categoryMatch = !!(p.badges || []).includes("hit");
            else if (state.category === "sale") categoryMatch = !!(p.badges || []).includes("sale");
            else if (state.category === "limited") categoryMatch = !!(p.badges || []).includes("limited");
            else categoryMatch = p.category_slug === state.category;
        }}
        const sizeMatch = !size || (p.sizes_list || []).includes(size);
        const priceMatch = (!min || Number(p.price || 0) >= min) && (!max || Number(p.price || 0) <= max);
        const favoriteMatch = !state.favoritesOnly || !!p.favorite;
        return textMatch && categoryMatch && sizeMatch && priceMatch && favoriteMatch;
    }});
    renderProducts();
}}

function badgeHtml(product) {{
    return (product.badges || []).map(key => {{
        const txt = (BADGE_LABELS[lang] || BADGE_LABELS.ru)[key] || key;
        return `<span class="badge ${{esc(key)}}">${{esc(txt)}}</span>`;
    }}).join("");
}}

function stockText(product) {{
    const stock = Number(product.stock_qty || 0);
    if (stock <= 0) return `<div class="stock out">${{esc(L.out)}}</div>`;
    if (stock <= 3) return `<div class="stock low">${{esc(L.left)}} ${{stock}} шт</div>`;
    return `<div class="stock">${{esc(L.left)}} ${{stock}} шт</div>`;
}}

function renderProducts() {{
    const grid = document.getElementById("productGrid");
    grid.innerHTML = "";
    if (!state.filtered.length) {{
        grid.innerHTML = `<div class="card"><div class="card-inner">${{esc(L.empty)}}</div></div>`;
        return;
    }}

    state.filtered.forEach(product => {{
        let activeSize = (product.sizes_list || [])[0] || "";

        const card = document.createElement("div");
        card.className = "card";

        const photoHtml = product.photo_url
            ? `<img src="${{esc(product.photo_url)}}" alt="${{esc(product.title)}}">`
            : `<div style="padding:18px;color:#8b7869;font-weight:700">No photo</div>`;

        const oldPriceHtml = Number(product.old_price || 0) > 0
            ? `<div class="old-price">${{fmtSum(product.old_price)}}</div>`
            : "";

        const discountHtml = Number(product.discount_percent || 0) > 0
            ? `<div class="discount-pill">-${{product.discount_percent}}%</div>`
            : "";

        const sizesHtml = (product.sizes_list || []).map((size, idx) =>
            `<button class="size-btn${{idx === 0 ? " active" : ""}}" data-size="${{esc(size)}}">${{esc(size)}}</button>`
        ).join("");

        card.innerHTML = `
            <div class="card-inner">
                <div class="photo-wrap">
                    <div class="badges">${{badgeHtml(product)}}</div>
                    <button class="fav${{product.favorite ? " active" : ""}}" title="favorite">${{product.favorite ? "❤️" : "🤍"}}</button>
                    <div class="photo">${{photoHtml}}</div>
                </div>
                <div class="title">${{esc(product.title)}}</div>
                <div class="desc">${{esc(product.description || "")}}</div>
                <div class="price-row">
                    <div class="price">${{fmtSum(product.price)}}</div>
                    ${{oldPriceHtml}}
                    ${{discountHtml}}
                </div>
                ${{stockText(product)}}
                <div class="sizes">${{sizesHtml}}</div>
                <div class="actions">
                    <button class="btn-dark addBtn" ${{product.can_buy ? "" : "disabled"}}>${{esc(L.add)}}</button>
                    <button class="btn-gold buyBtn" ${{product.can_buy ? "" : "disabled"}}>${{esc(L.buy)}}</button>
                </div>
            </div>
        `;

        card.querySelectorAll(".size-btn").forEach(btn => {{
            btn.onclick = () => {{
                card.querySelectorAll(".size-btn").forEach(x => x.classList.remove("active"));
                btn.classList.add("active");
                activeSize = btn.dataset.size || "";
            }};
        }});

        card.querySelector(".fav").onclick = async () => {{
            if (!userId) return notice("Telegram user not found");
            const res = await api(`/api/shop/favorite/toggle`, {{
                method: "POST",
                headers: {{ "Content-Type": "application/json" }},
                body: JSON.stringify({{ user_id: userId, product_id: product.id }})
            }});
            product.favorite = !!res.favorite;
            notice(product.favorite ? L.favoriteAdded : L.favoriteRemoved);
            renderProducts();
        }};

        card.querySelector(".addBtn").onclick = () => {{
            sendData({{ action: "add_to_cart", product_id: product.id, qty: 1, size: activeSize }});
            notice(L.added);
            setTimeout(loadCart, 500);
        }};

        card.querySelector(".buyBtn").onclick = () => {{
            sendData({{ action: "buy_now", product_id: product.id, qty: 1, size: activeSize }});
            notice(L.checkout);
            setTimeout(loadCart, 500);
        }};

        grid.appendChild(card);
    }});
}}

function renderCart() {{
    const box = document.getElementById("cartList");
    const qtyEl = document.getElementById("sumQty");
    const amountEl = document.getElementById("sumAmount");
    box.innerHTML = "";

    let totalQty = 0;
    let totalAmount = 0;

    (state.cart || []).forEach(item => {{
        totalQty += Number(item.qty || 0);
        totalAmount += Number(item.subtotal || 0);

        const div = document.createElement("div");
        div.className = "cart-item";
        div.innerHTML = `
            <div><b>${{esc(item.product_name)}}</b></div>
            <div class="mini-note">${{item.size ? esc(item.size) + " | " : ""}}${{item.qty}} × ${{fmtSum(item.price)}}</div>
            <div style="display:flex;justify-content:space-between;align-items:center;margin-top:8px;gap:10px">
                <b>${{fmtSum(item.subtotal)}}</b>
                <button class="btn-ghost">Удалить</button>
            </div>
        `;
        div.querySelector("button").onclick = () => {{
            sendData({{ action: "remove_from_cart", cart_id: item.cart_id }});
            notice(L.removed);
            setTimeout(loadCart, 450);
        }};
        box.appendChild(div);
    }});

    if (!state.cart.length) {{
        box.innerHTML = `<div class="empty">${{esc(L.cartEmpty)}}</div>`;
    }}

    qtyEl.textContent = String(totalQty);
    amountEl.textContent = fmtSum(totalAmount);
}}

function renderReviews() {{
    const box = document.getElementById("reviewsBox");
    box.innerHTML = "";
    if (!state.reviews.length) {{
        box.innerHTML = `<div class="empty">${{esc(L.reviewsEmpty)}}</div>`;
        return;
    }}
    state.reviews.forEach(r => {{
        const div = document.createElement("div");
        div.className = "review";
        div.innerHTML = `
            <div class="name">${{esc(r.customer_name || "Client")}} — ${{"⭐".repeat(Math.max(1, Math.min(5, Number(r.rating || 5))))}}</div>
            <div class="text">${{esc(r.text || "")}}</div>
        `;
        box.appendChild(div);
    }});
}}

document.getElementById("searchInput").addEventListener("input", e => {{
    state.search = e.target.value || "";
    applyFilters();
}});
document.getElementById("sizeFilter").addEventListener("change", e => {{
    state.size = e.target.value || "";
    applyFilters();
}});
document.getElementById("priceMin").addEventListener("input", e => {{
    state.priceMin = e.target.value || "";
    applyFilters();
}});
document.getElementById("priceMax").addEventListener("input", e => {{
    state.priceMax = e.target.value || "";
    applyFilters();
}});
document.getElementById("favOnlyBtn").onclick = () => {{
    state.favoritesOnly = !state.favoritesOnly;
    document.getElementById("favOnlyBtn").textContent = state.favoritesOnly ? "❤️ On" : "❤️ Избранное";
    applyFilters();
}};
document.getElementById("checkoutBtn").onclick = () => sendData({{ action: "checkout" }});
document.getElementById("clearBtn").onclick = () => {{
    sendData({{ action: "clear_cart" }});
    notice(L.cleared);
    setTimeout(loadCart, 450);
}};

createPetals();
buildCategoryFilters();
loadProducts();
loadCart();
loadReviews();
</script>
</body>
</html>
"""


# ============================================================
# ADMIN HTML
# ============================================================

def admin_allowed(request: web.Request) -> bool:
    token = request.query.get("token", "").strip()
    return bool(ADMIN_PANEL_TOKEN) and token == ADMIN_PANEL_TOKEN


def admin_token(request: web.Request) -> str:
    return request.query.get("token", "").strip()


def admin_url(path: str, token: str) -> str:
    sep = "&" if "?" in path else "?"
    return f"{path}{sep}token={escape(token)}"


def bool_from_form(value: Any) -> int:
    if str(value).lower() in {"1", "true", "on", "yes"}:
        return 1
    return 0


def admin_template(title: str, body: str, token: str) -> str:
    nav = f"""
    <div class="nav">
        <a href="{admin_url('/admin', token)}">Dashboard</a>
        <a href="{admin_url('/admin/orders', token)}">Orders</a>
        <a href="{admin_url('/admin/products', token)}">Products</a>
        <a href="{admin_url('/admin/reviews', token)}">Reviews</a>
    </div>
    """
    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{escape(title)}</title>
<style>
:root {{
    --bg1:#fffaf4;
    --bg2:#fcf1f7;
    --gold:#c7a865;
    --gold2:#9f7841;
    --text:#221912;
    --muted:#7f6d61;
    --line:#eadfcf;
}}
* {{ box-sizing:border-box; }}
body {{
    margin:0;
    font-family:Inter,Arial,sans-serif;
    color:var(--text);
    background:linear-gradient(135deg,var(--bg1),var(--bg2),#fffaf2);
}}
.wrap {{ max-width:1280px; margin:0 auto; padding:20px 14px 40px; }}
.top {{
    background:rgba(255,255,255,.78);
    border-radius:28px;
    padding:20px;
    box-shadow:0 18px 48px rgba(90,65,30,.06);
}}
.brand {{
    font-size:34px;
    font-weight:900;
    letter-spacing:.14em;
    text-transform:uppercase;
    background:linear-gradient(180deg,#f7ecc8,var(--gold),var(--gold2));
    -webkit-background-clip:text;
    -webkit-text-fill-color:transparent;
}}
.nav {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:12px; }}
.nav a {{
    text-decoration:none;
    color:#fff;
    background:linear-gradient(180deg,var(--gold),var(--gold2));
    padding:10px 14px;
    border-radius:999px;
    font-weight:800;
}}
.card {{
    background:rgba(255,255,255,.82);
    border-radius:24px;
    padding:18px;
    margin-top:16px;
    box-shadow:0 14px 34px rgba(90,65,30,.05);
}}
.stats {{
    display:grid;
    grid-template-columns:repeat(4,minmax(0,1fr));
    gap:12px;
}}
.stat {{
    background:#fff;
    border:1px solid var(--line);
    border-radius:18px;
    padding:14px;
}}
.stat .label {{ font-size:12px; color:var(--muted); }}
.stat .value {{ font-size:28px; font-weight:900; margin-top:8px; }}
table {{ width:100%; border-collapse:collapse; background:#fff; border-radius:18px; overflow:hidden; }}
th,td {{ padding:10px 8px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; font-size:14px; }}
th {{ background:#faf4ea; }}
input,textarea,select {{
    width:100%;
    border:1px solid var(--line);
    border-radius:12px;
    padding:10px 12px;
    font:inherit;
    background:#fff;
}}
textarea {{ min-height:110px; resize:vertical; }}
.grid2 {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; }}
.grid3 {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:14px; }}
.actions {{ display:flex; gap:10px; flex-wrap:wrap; margin-top:14px; }}
.btn {{
    border:none;
    cursor:pointer;
    padding:12px 16px;
    border-radius:14px;
    font-weight:900;
}}
.btn.gold {{ color:#fff; background:linear-gradient(180deg,var(--gold),var(--gold2)); }}
.btn.dark {{ color:#fff; background:#19130f; }}
.btn.light {{ background:#fff; border:1px solid var(--line); }}
.badge {{
    display:inline-block;
    padding:4px 8px;
    border-radius:999px;
    font-size:12px;
    font-weight:800;
    background:#f5ede1;
}}
.notice {{
    padding:12px 14px;
    border-radius:14px;
    background:#f7f1e6;
    border:1px solid #eadcc5;
    margin-bottom:14px;
    font-weight:700;
}}
@media(max-width:980px) {{
    .stats {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    .grid2,.grid3 {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class="wrap">
    <div class="top">
        <div class="brand">{escape(SHOP_BRAND)} ADMIN</div>
        {nav}
    </div>
    {body}
</div>
</body>
</html>
"""


# ============================================================
# USER HANDLERS
# ============================================================

@user_router.message(CommandStart())
async def start_handler(message: Message) -> None:
    upsert_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    await message.answer(t(message.from_user.id, "welcome"), reply_markup=user_main_menu(message.from_user.id))
    await message.answer(t(message.from_user.id, "main_menu_hint"))


@user_router.message(F.text.in_([TEXTS["ru"]["menu_shop"], TEXTS["uz"]["menu_shop"]]))
async def shop_menu_handler(message: Message) -> None:
    await message.answer(
        t(message.from_user.id, "shop_opened"),
        reply_markup=user_main_menu(message.from_user.id),
    )


@user_router.message(F.text.in_([TEXTS["ru"]["menu_lang"], TEXTS["uz"]["menu_lang"]]))
async def choose_lang_handler(message: Message) -> None:
    await message.answer(t(message.from_user.id, "choose_lang"), reply_markup=language_keyboard())


@user_router.callback_query(F.data.startswith("lang:"))
async def set_lang_callback(callback: CallbackQuery) -> None:
    lang = callback.data.split(":")[-1]
    set_user_lang(callback.from_user.id, lang)
    await callback.message.answer(t(callback.from_user.id, "lang_updated"), reply_markup=user_main_menu(callback.from_user.id))
    await callback.answer()


@user_router.message(F.text.in_([TEXTS["ru"]["menu_contacts"], TEXTS["uz"]["menu_contacts"]]))
async def contacts_handler(message: Message) -> None:
    text = (
        f"{t(message.from_user.id, 'contacts_title')}\n\n"
        f"<b>Phone:</b> {escape(MANAGER_PHONE)}\n"
        f"<b>Telegram:</b> {escape(MANAGER_TG)}\n"
        f"<b>Channel:</b> {escape(CHANNEL_LINK or '—')}\n"
        f"<b>Instagram:</b> {escape(INSTAGRAM_LINK or '—')}\n"
        f"<b>YouTube:</b> {escape(YOUTUBE_LINK or '—')}"
    )
    await message.answer(text, reply_markup=user_main_menu(message.from_user.id))


# ============================================================
# CART HANDLERS
# ============================================================

@cart_router.message(F.text.in_([TEXTS["ru"]["menu_cart"], TEXTS["uz"]["menu_cart"]]))
async def cart_handler(message: Message) -> None:
    data = cart_api(message.from_user.id, get_user_lang(message.from_user.id))
    if not data["items"]:
        await message.answer(t(message.from_user.id, "cart_empty"))
        return

    lines = ["🛒 <b>Корзина</b>"]
    for idx, item in enumerate(data["items"], start=1):
        size = f" | {escape(item['size'])}" if item["size"] else ""
        lines.append(
            f"{idx}. <b>{escape(item['product_name'])}</b>{size} — {item['qty']} × {fmt_sum(item['price'])}"
        )
    lines.append("")
    lines.append(f"<b>Всего товаров:</b> {data['total_qty']}")
    lines.append(f"<b>Сумма:</b> {fmt_sum(data['total_amount'])}")

    await message.answer("\n".join(lines))


@cart_router.message(F.web_app_data)
async def webapp_data_handler(message: Message, state: FSMContext) -> None:
    upsert_user(message.from_user.id, message.from_user.username, message.from_user.full_name)

    try:
        payload = json.loads(message.web_app_data.data)
    except Exception:
        await message.answer("Bad web app payload")
        return

    action = str(payload.get("action") or "").strip()

    if action == "add_to_cart":
        ok, key = add_to_cart(
            message.from_user.id,
            safe_int(payload.get("product_id")),
            max(1, safe_int(payload.get("qty"), 1)),
            str(payload.get("size") or "")
        )
        if ok:
            await message.answer(t(message.from_user.id, key))
        else:
            await message.answer(t(message.from_user.id, "stock_out") if key == "stock_out" else "Ошибка")
        return

    if action == "remove_from_cart":
        ok = remove_cart_item(message.from_user.id, safe_int(payload.get("cart_id")))
        await message.answer(t(message.from_user.id, "cart_removed") if ok else "Not found")
        return

    if action == "clear_cart":
        clear_cart(message.from_user.id)
        await message.answer(t(message.from_user.id, "cart_cleared"))
        return

    if action == "checkout":
        if not get_cart_rows(message.from_user.id):
            await message.answer(t(message.from_user.id, "cart_empty"))
            return
        await state.clear()
        await state.set_state(CheckoutState.customer_name)
        await message.answer(
            f"{t(message.from_user.id, 'checkout_intro')}\n\n{t(message.from_user.id, 'checkout_name')}",
            reply_markup=cancel_keyboard(message.from_user.id),
        )
        return

    if action == "buy_now":
        ok, _ = add_to_cart(
            message.from_user.id,
            safe_int(payload.get("product_id")),
            max(1, safe_int(payload.get("qty"), 1)),
            str(payload.get("size") or "")
        )
        if not ok:
            await message.answer(t(message.from_user.id, "stock_out"))
            return
        await state.clear()
        await state.set_state(CheckoutState.customer_name)
        await message.answer(
            f"{t(message.from_user.id, 'checkout_intro')}\n\n{t(message.from_user.id, 'checkout_name')}",
            reply_markup=cancel_keyboard(message.from_user.id),
        )
        return


# ============================================================
# CHECKOUT HANDLERS
# ============================================================

async def maybe_cancel(message: Message, state: FSMContext) -> bool:
    if (message.text or "").strip() == t(message.from_user.id, "cancel"):
        await state.clear()
        await message.answer("Отменено.", reply_markup=user_main_menu(message.from_user.id))
        return True
    return False


@checkout_router.message(CheckoutState.customer_name)
async def checkout_name_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    name = (message.text or "").strip()
    if not name:
        await message.answer(t(message.from_user.id, "checkout_name"))
        return

    await state.update_data(customer_name=name)
    await state.set_state(CheckoutState.customer_phone)
    await message.answer(t(message.from_user.id, "checkout_phone"))


@checkout_router.message(CheckoutState.customer_phone)
async def checkout_phone_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    phone = normalize_phone(message.text or "")
    if not is_valid_phone(phone):
        await message.answer(t(message.from_user.id, "checkout_invalid_phone"))
        return

    await state.update_data(customer_phone=phone)
    await state.set_state(CheckoutState.delivery_method)
    await message.answer(t(message.from_user.id, "checkout_delivery"), reply_markup=delivery_keyboard(message.from_user.id))


@checkout_router.message(CheckoutState.delivery_method)
async def checkout_delivery_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    method = delivery_method_from_label(message.from_user.id, message.text or "")
    if not method:
        await message.answer(t(message.from_user.id, "checkout_delivery"), reply_markup=delivery_keyboard(message.from_user.id))
        return

    await state.update_data(delivery_method=method)
    await state.set_state(CheckoutState.city)
    await message.answer(t(message.from_user.id, "checkout_city"), reply_markup=cancel_keyboard(message.from_user.id))


@checkout_router.message(CheckoutState.city)
async def checkout_city_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    city = (message.text or "").strip()
    if not city:
        await message.answer(t(message.from_user.id, "checkout_city"))
        return

    await state.update_data(city=city)
    await state.set_state(CheckoutState.delivery_address)
    await message.answer(t(message.from_user.id, "checkout_address"))


@checkout_router.message(CheckoutState.delivery_address)
async def checkout_address_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    addr = (message.text or "").strip()
    if not addr:
        await message.answer(t(message.from_user.id, "checkout_address"))
        return

    await state.update_data(delivery_address=addr)
    await state.set_state(CheckoutState.payment_method)
    await message.answer(t(message.from_user.id, "checkout_payment"), reply_markup=payment_keyboard(message.from_user.id))


@checkout_router.message(CheckoutState.payment_method)
async def checkout_payment_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    method = payment_method_from_label(message.from_user.id, message.text or "")
    if not method:
        await message.answer(t(message.from_user.id, "checkout_payment"), reply_markup=payment_keyboard(message.from_user.id))
        return

    await state.update_data(payment_method=method)
    await state.set_state(CheckoutState.comment)
    await message.answer(t(message.from_user.id, "checkout_comment"), reply_markup=skip_keyboard(message.from_user.id))


@checkout_router.message(CheckoutState.comment)
async def checkout_comment_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    comment = (message.text or "").strip()
    if comment == t(message.from_user.id, "skip"):
        comment = ""

    await state.update_data(comment=comment)

    data = await state.get_data()
    cart_rows = get_cart_rows(message.from_user.id)
    if not cart_rows:
        await state.clear()
        await message.answer(t(message.from_user.id, "cart_empty"), reply_markup=user_main_menu(message.from_user.id))
        return

    lines = [t(message.from_user.id, "checkout_summary"), ""]
    lines.append(f"<b>Имя:</b> {escape(data.get('customer_name'))}")
    lines.append(f"<b>Телефон:</b> {escape(data.get('customer_phone'))}")
    lines.append(f"<b>Город:</b> {escape(data.get('city'))}")
    lines.append(f"<b>Адрес:</b> {escape(data.get('delivery_address'))}")
    lines.append(f"<b>Доставка:</b> {delivery_label(message.from_user.id, data.get('delivery_method', 'pickup'))}")
    lines.append(f"<b>Оплата:</b> {escape(data.get('payment_method', 'cash'))}")
    lines.append("")
    lines.append("<b>Товары:</b>")

    total_qty, total_amount = cart_totals(message.from_user.id)
    for idx, row in enumerate(cart_rows, start=1):
        size = f" | {escape(row['size'])}" if row["size"] else ""
        lines.append(f"{idx}. {escape(row['title_uz'] if get_user_lang(message.from_user.id) == 'uz' else row['title_ru'])}{size} — {row['qty']} × {fmt_sum(row['price'])}")

    lines.append("")
    lines.append(f"<b>Количество:</b> {total_qty}")
    lines.append(f"<b>Сумма:</b> {fmt_sum(total_amount)}")
    lines.append("")
    lines.append(t(message.from_user.id, "checkout_confirm_hint"))

    await state.set_state(CheckoutState.confirm)
    await message.answer("\n".join(lines), reply_markup=yes_no_keyboard(message.from_user.id))


@checkout_router.message(CheckoutState.confirm)
async def checkout_confirm_handler(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()

    if text == t(message.from_user.id, "no"):
        await state.clear()
        await message.answer("Оформление отменено.", reply_markup=user_main_menu(message.from_user.id))
        return

    if text != t(message.from_user.id, "yes"):
        await message.answer(t(message.from_user.id, "checkout_confirm_hint"))
        return

    data = await state.get_data()
    order_id = create_order(message.from_user.id, message.from_user.username or "", data, source="telegram")
    update_user_profile(message.from_user.id, data.get("customer_phone", ""), data.get("city", ""))
    await state.clear()

    order = get_order(order_id)
    await message.answer(
        f"{t(message.from_user.id, 'checkout_done')}\n\n"
        f"<b>Номер заказа:</b> {escape(order['order_number'])}\n"
        f"<b>Сумма:</b> {fmt_sum(order['total_amount'])}",
        reply_markup=user_main_menu(message.from_user.id),
    )
    await notify_admins_new_order(order_id)
    await notify_user_order_status(order_id, "new")
    await notify_admins_low_stock()


# ============================================================
# ORDERS HANDLERS
# ============================================================

@orders_router.message(F.text.in_([TEXTS["ru"]["menu_orders"], TEXTS["uz"]["menu_orders"]]))
async def user_orders_handler(message: Message) -> None:
    rows = get_orders_for_user(message.from_user.id)
    if not rows:
        await message.answer("Заказов пока нет.")
        return

    for row in rows:
        await message.answer(
            f"<b>Заказ {escape(row['order_number'] or row['id'])}</b>\n"
            f"Дата: {escape(row['created_at'])}\n"
            f"Статус: {status_label(message.from_user.id, row['status'])}\n"
            f"Оплата: {payment_status_label(message.from_user.id, row['payment_status'])}\n"
            f"Доставка: {delivery_label(message.from_user.id, row['delivery_method'])}\n"
            f"Сумма: {fmt_sum(row['total_amount'])}\n\n"
            f"{render_items_for_message(row['items_json'])}"
        )


# ============================================================
# REVIEWS HANDLERS
# ============================================================

@reviews_router.message(F.text.in_([TEXTS["ru"]["menu_reviews"], TEXTS["uz"]["menu_reviews"]]))
async def reviews_list_handler(message: Message) -> None:
    conn = get_db()
    rows = conn.execute("SELECT * FROM reviews WHERE is_published = 1 ORDER BY id DESC LIMIT 20").fetchall()
    conn.close()

    if not rows:
        await message.answer(t(message.from_user.id, "reviews_empty"))
        return

    for row in rows:
        stars = "⭐" * max(1, min(5, safe_int(row["rating"], 5)))
        await message.answer(
            f"{stars}\n"
            f"<b>{escape(row['customer_name'] or 'Client')}</b>\n"
            f"{escape(row['text'])}"
        )


@reviews_router.message(F.text.in_([TEXTS["ru"]["menu_leave_review"], TEXTS["uz"]["menu_leave_review"]]))
async def leave_review_start_handler(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(ReviewState.rating)
    await message.answer(t(message.from_user.id, "review_ask_rating"), reply_markup=cancel_keyboard(message.from_user.id))


@reviews_router.message(ReviewState.rating)
async def review_rating_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    rating = safe_int((message.text or "").strip(), 0)
    if rating not in {1, 2, 3, 4, 5}:
        await message.answer(t(message.from_user.id, "review_bad_rating"))
        return

    await state.update_data(rating=rating)
    await state.set_state(ReviewState.text)
    await message.answer(t(message.from_user.id, "review_ask_text"), reply_markup=cancel_keyboard(message.from_user.id))


@reviews_router.message(ReviewState.text)
async def review_text_handler(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return

    text = (message.text or "").strip()
    if not text:
        await message.answer(t(message.from_user.id, "review_ask_text"))
        return

    data = await state.get_data()
    add_review(
        message.from_user.id,
        message.from_user.username or "",
        message.from_user.full_name or "",
        safe_int(data.get("rating"), 5),
        text,
    )
    await state.clear()
    await message.answer(t(message.from_user.id, "review_sent"), reply_markup=user_main_menu(message.from_user.id))


# ============================================================
# ADMIN HANDLERS
# ============================================================

@admin_router.message(F.text.in_([TEXTS["ru"]["menu_admin"], TEXTS["uz"]["menu_admin"]]))
async def admin_open_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return
    await message.answer(t(message.from_user.id, "admin_title"), reply_markup=admin_menu(message.from_user.id))


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_back"], TEXTS["uz"]["admin_back"]]))
async def admin_back_handler(message: Message) -> None:
    await message.answer(t(message.from_user.id, "main_menu_hint"), reply_markup=user_main_menu(message.from_user.id))


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_stats"], TEXTS["uz"]["admin_stats"]]))
async def admin_stats_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    stats = get_admin_stats()
    await message.answer(
        f"📊 <b>Статистика</b>\n\n"
        f"Заказы: <b>{stats['orders']}</b>\n"
        f"Новые заказы: <b>{stats['new_orders']}</b>\n"
        f"Пользователи: <b>{stats['users']}</b>\n"
        f"Товары: <b>{stats['products']}</b>\n"
        f"Опубликованные отзывы: <b>{stats['reviews']}</b>\n"
        f"Мало на складе: <b>{stats['low_stock']}</b>"
    )


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_stock"], TEXTS["uz"]["admin_stock"]]))
async def admin_stock_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    rows = low_stock_products()
    if not rows:
        await message.answer("На складе всё нормально.")
        return

    lines = ["📉 <b>Остатки</b>"]
    for row in rows[:40]:
        lines.append(f"#{row['id']} {escape(row['title_ru'])} — {row['stock_qty']} шт")
    lines.append("\nКоманда: /set_stock PRODUCT_ID QTY")
    await message.answer("\n".join(lines))


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_products"], TEXTS["uz"]["admin_products"]]))
async def admin_products_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    rows = get_products(published_only=False)[:40]
    if not rows:
        await message.answer("Товаров пока нет.")
        return

    for row in rows:
        await message.answer(
            f"📦 <b>Товар #{row['id']}</b>\n"
            f"RU: {escape(row['title_ru'])}\n"
            f"UZ: {escape(row['title_uz'])}\n"
            f"Цена: {fmt_sum(row['price'])}\n"
            f"Старая цена: {fmt_sum(row['old_price']) if safe_int(row['old_price']) else '—'}\n"
            f"Остаток: {row['stock_qty']}\n"
            f"Размеры: {escape(row['sizes'] or '—')}\n"
            f"Категория: {escape(row['category_slug'])}\n"
            f"Опубликован: {'Да' if safe_int(row['is_published']) else 'Нет'}\n\n"
            f"Команды:\n"
            f"/set_price {row['id']} 399000\n"
            f"/set_stock {row['id']} 7\n"
            f"/toggle_product {row['id']}"
        )


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_orders"], TEXTS["uz"]["admin_orders"]]))
async def admin_orders_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    conn = get_db()
    rows = conn.execute("SELECT * FROM orders ORDER BY id DESC LIMIT 25").fetchall()
    conn.close()

    if not rows:
        await message.answer("Пока заказов нет.")
        return

    for row in rows:
        await message.answer(
            f"📦 <b>Заказ {escape(row['order_number'] or row['id'])}</b>\n"
            f"{escape(row['customer_name'])} | {escape(row['customer_phone'])}\n"
            f"{status_label('ru', row['status'])} | {payment_status_label('ru', row['payment_status'])}\n"
            f"{delivery_label('ru', row['delivery_method'])}\n"
            f"{fmt_sum(row['total_amount'])}\n\n"
            f"{render_items_for_message(row['items_json'])}\n\n"
            f"Команды:\n"
            f"/set_order_status {row['id']} confirmed\n"
            f"/set_payment_status {row['id']} paid"
        )


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_reviews"], TEXTS["uz"]["admin_reviews"]]))
async def admin_reviews_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    conn = get_db()
    rows = conn.execute("SELECT * FROM reviews ORDER BY id DESC LIMIT 25").fetchall()
    conn.close()

    if not rows:
        await message.answer("Отзывов пока нет.")
        return

    for row in rows:
        stars = "⭐" * max(1, min(5, safe_int(row["rating"], 5)))
        await message.answer(
            f"⭐ <b>Отзыв #{row['id']}</b>\n"
            f"Имя: {escape(row['customer_name'])}\n"
            f"Оценка: {stars}\n"
            f"Статус: {'Опубликован' if safe_int(row['is_published']) else 'На модерации'}\n\n"
            f"{escape(row['text'])}\n\n"
            f"Команда: /publish_review {row['id']}"
        )


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_customers"], TEXTS["uz"]["admin_customers"]]))
async def admin_customers_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    conn = get_db()
    rows = conn.execute(
        """
        SELECT customer_phone, customer_name, COUNT(*) AS orders_count, SUM(total_amount) AS total_spent
        FROM orders
        GROUP BY customer_phone
        ORDER BY orders_count DESC, total_spent DESC
        LIMIT 30
        """
    ).fetchall()
    conn.close()

    if not rows:
        await message.answer("Клиентов пока нет.")
        return

    lines = ["👥 <b>Клиенты</b>"]
    for row in rows:
        lines.append(
            f"• {escape(row['customer_name'] or '—')} | {escape(row['customer_phone'] or '—')} | "
            f"заказов: {row['orders_count']} | сумма: {fmt_sum(row['total_spent'] or 0)}"
        )
    await message.answer("\n".join(lines))


@admin_router.message(F.text.in_([TEXTS["ru"]["admin_report"], TEXTS["uz"]["admin_report"]]))
async def admin_report_hint_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return
    await message.answer("Команда: /report 2026 03")


@admin_router.message(Command("set_stock"))
async def admin_set_stock(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    parts = (message.text or "").split()
    if len(parts) != 3:
        await message.answer("Используй: /set_stock PRODUCT_ID QTY")
        return

    product_id = safe_int(parts[1])
    qty = max(0, safe_int(parts[2]))
    conn = get_db()
    conn.execute("UPDATE products SET stock_qty = ?, updated_at = ? WHERE id = ?", (qty, utc_now_iso(), product_id))
    conn.commit()
    conn.close()

    await message.answer(f"Остаток товара #{product_id} обновлён: {qty} шт")


@admin_router.message(Command("set_price"))
async def admin_set_price(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    parts = (message.text or "").split()
    if len(parts) != 3:
        await message.answer("Используй: /set_price PRODUCT_ID PRICE")
        return

    product_id = safe_int(parts[1])
    price = max(0, safe_int(parts[2]))
    conn = get_db()
    conn.execute("UPDATE products SET price = ?, updated_at = ? WHERE id = ?", (price, utc_now_iso(), product_id))
    conn.commit()
    conn.close()

    await message.answer(f"Цена товара #{product_id} обновлена: {fmt_sum(price)}")


@admin_router.message(Command("toggle_product"))
async def admin_toggle_product(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return

    parts = (message.text or "").split()
    if len(parts) != 2:
        await message.answer("Используй: /toggle_product PRODUCT_ID")
        return

    product_id = safe_int(parts[1])
    conn = get_db()
    row = conn.execute("SELECT is_published FROM products WHERE id = ?", (product_id,)).fetchone()
    if not row:
        conn.close()
        await message.answer("Товар не найден.")
        return

    new_value = 0 if safe_int(row["is_published"]) else 1
    conn.execute("UPDATE products SET is_published = ?, updated_at = ? WHERE id = ?", (new_value, utc_now_iso(), product_id))
    conn.commit()
    conn.close()

    await message.answer(f"Товар #{product_id} теперь {'опубликован' if new_value else 'скрыт'}")


# ============================================================
# EXTRA DB / NOTIFY LOGS / MONTHLY LOG
# ============================================================

def init_extra_db() -> None:
    conn = get_db()
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS order_notify_state (
            order_id INTEGER PRIMARY KEY,
            last_status TEXT NOT NULL DEFAULT '',
            last_payment_status TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS monthly_report_log (
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            sent_at TEXT NOT NULL,
            PRIMARY KEY (year, month)
        );
        """
    )
    conn.commit()
    conn.close()


# ============================================================
# REVIEW FSM
# ============================================================

@reviews_router.message(F.text.in_([TEXTS["ru"]["menu_leave_review"], TEXTS["uz"]["menu_leave_review"]]))
async def leave_review_start(message: Message, state: FSMContext) -> None:
    upsert_user(message.from_user.id, message.from_user.username, message.from_user.full_name)
    await state.clear()
    await state.set_state(ReviewState.rating)
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="5"), KeyboardButton(text="4"), KeyboardButton(text="3")],
            [KeyboardButton(text="2"), KeyboardButton(text="1")],
            [KeyboardButton(text=t(message.from_user.id, "cancel"))],
        ],
        resize_keyboard=True,
    )
    await message.answer("Оцените магазин от 1 до 5.", reply_markup=kb)


@reviews_router.message(ReviewState.rating)
async def leave_review_rating(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return
    rating = safe_int((message.text or "").strip(), 0)
    if rating < 1 or rating > 5:
        await message.answer("Отправьте число от 1 до 5.")
        return
    await state.update_data(rating=rating)
    await state.set_state(ReviewState.text)
    await message.answer("Напишите текст отзыва.", reply_markup=cancel_keyboard(message.from_user.id))


@reviews_router.message(ReviewState.text)
async def leave_review_text(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return
    text = (message.text or "").strip()
    if len(text) < 3:
        await message.answer("Текст отзыва слишком короткий.")
        return
    data = await state.get_data()
    now = utc_now_iso()
    conn = get_db()
    conn.execute(
        """
        INSERT INTO reviews (
            user_id, username, customer_name, rating, text, is_published, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, 0, ?, ?)
        """,
        (
            message.from_user.id,
            message.from_user.username or "",
            message.from_user.full_name or "",
            safe_int(data.get("rating"), 5),
            text,
            now,
            now,
        ),
    )
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(t(message.from_user.id, "review_sent"), reply_markup=user_main_menu(message.from_user.id))
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id,
                f"⭐ <b>Новый отзыв на модерации</b>\n"
                f"От: {html.escape(message.from_user.full_name or 'Client')}\n"
                f"Оценка: {'⭐' * max(1, min(5, safe_int(data.get('rating'), 5)))}\n\n"
                f"{html.escape(text)}"
            )
        except Exception:
            logger.exception("Failed to notify admin about review")


# ============================================================
# ORDER / PAYMENT NOTIFICATIONS TO CLIENT
# ============================================================

def order_client_text(lang: str, order_id: int, status: str) -> Optional[str]:
    lang = ensure_lang(lang)
    status_map = {
        "new": {
            "ru": f"✅ Заказ <b>#{order_id}</b> принят и создан.",
            "uz": f"✅ <b>#{order_id}</b> buyurtma qabul qilindi va yaratildi.",
        },
        "confirmed": {
            "ru": f"📋 Заказ <b>#{order_id}</b> подтверждён.",
            "uz": f"📋 <b>#{order_id}</b> buyurtma tasdiqlandi.",
        },
        "paid": {
            "ru": f"💳 Заказ <b>#{order_id}</b> оплачен.",
            "uz": f"💳 <b>#{order_id}</b> buyurtma to'landi.",
        },
        "sent": {
            "ru": f"🚚 Заказ <b>#{order_id}</b> отправлен.",
            "uz": f"🚚 <b>#{order_id}</b> buyurtma yuborildi.",
        },
        "delivered": {
            "ru": f"🎉 Заказ <b>#{order_id}</b> доставлен. Спасибо за покупку.",
            "uz": f"🎉 <b>#{order_id}</b> buyurtma yetkazildi. Xarid uchun rahmat.",
        },
        "cancelled": {
            "ru": f"❌ Заказ <b>#{order_id}</b> отменён.",
            "uz": f"❌ <b>#{order_id}</b> buyurtma bekor qilindi.",
        },
    }
    return status_map.get(status, {}).get(lang)


def payment_client_text(lang: str, order_id: int, status: str) -> Optional[str]:
    lang = ensure_lang(lang)
    status_map = {
        "pending": {
            "ru": f"⏳ Оплата заказа <b>#{order_id}</b> ожидается.",
            "uz": f"⏳ <b>#{order_id}</b> buyurtma to'lovi kutilmoqda.",
        },
        "paid": {
            "ru": f"✅ Оплата заказа <b>#{order_id}</b> подтверждена.",
            "uz": f"✅ <b>#{order_id}</b> buyurtma to'lovi tasdiqlandi.",
        },
        "failed": {
            "ru": f"⚠️ Оплата заказа <b>#{order_id}</b> завершилась ошибкой.",
            "uz": f"⚠️ <b>#{order_id}</b> buyurtma to'lovida xatolik yuz berdi.",
        },
        "cancelled": {
            "ru": f"❌ Оплата заказа <b>#{order_id}</b> отменена.",
            "uz": f"❌ <b>#{order_id}</b> buyurtma to'lovi bekor qilindi.",
        },
        "refunded": {
            "ru": f"↩️ По заказу <b>#{order_id}</b> выполнен возврат.",
            "uz": f"↩️ <b>#{order_id}</b> buyurtma bo'yicha qaytarish bajarildi.",
        },
    }
    return status_map.get(status, {}).get(lang)


async def sync_order_notifications_once() -> None:
    conn = get_db()
    init_extra_db()
    rows = conn.execute("SELECT * FROM orders ORDER BY id DESC LIMIT 500").fetchall()
    for row in rows:
        state = conn.execute(
            "SELECT * FROM order_notify_state WHERE order_id = ?",
            (row["id"],),
        ).fetchone()

        user_id = safe_int(row["user_id"])
        lang = get_user_lang(user_id)
        created_at = row["created_at"] or ""
        recent_new_order = created_at[:16] == utc_now_iso()[:16] or created_at[:13] == utc_now_iso()[:13]

        if not state:
            # Для очень свежих заказов шлём "принят", старые просто синхронизируем.
            if recent_new_order:
                text = order_client_text(lang, row["id"], "new")
                if text:
                    try:
                        await bot.send_message(user_id, text)
                    except Exception:
                        logger.exception("Failed to send accepted notification to %s", user_id)
            conn.execute(
                """
                INSERT INTO order_notify_state (order_id, last_status, last_payment_status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (row["id"], row["status"], row["payment_status"], utc_now_iso(), utc_now_iso()),
            )
            conn.commit()
            continue

        last_status = state["last_status"] or ""
        last_payment_status = state["last_payment_status"] or ""

        if row["status"] != last_status:
            text = order_client_text(lang, row["id"], row["status"])
            if text:
                try:
                    await bot.send_message(
                        user_id,
                        text + f"\n\nСтатус: <b>{status_label(lang, row['status'])}</b>\nСумма: <b>{fmt_sum(row['total_amount'])}</b>",
                    )
                except Exception:
                    logger.exception("Failed to send order status notification to %s", user_id)

        if row["payment_status"] != last_payment_status:
            text = payment_client_text(lang, row["id"], row["payment_status"])
            if text:
                try:
                    await bot.send_message(
                        user_id,
                        text + f"\n\nОплата: <b>{payment_status_label(lang, row['payment_status'])}</b>",
                    )
                except Exception:
                    logger.exception("Failed to send payment status notification to %s", user_id)

        if row["status"] != last_status or row["payment_status"] != last_payment_status:
            conn.execute(
                """
                UPDATE order_notify_state
                SET last_status = ?, last_payment_status = ?, updated_at = ?
                WHERE order_id = ?
                """,
                (row["status"], row["payment_status"], utc_now_iso(), row["id"]),
            )
            conn.commit()

    conn.close()


async def order_notify_loop() -> None:
    while True:
        try:
            await sync_order_notifications_once()
        except Exception:
            logger.exception("order_notify_loop failed")
        await asyncio.sleep(20)


# ============================================================
# MONTHLY AUTO REPORT
# ============================================================

def previous_month_year_month(now_dt: datetime) -> tuple[int, int]:
    y = now_dt.year
    m = now_dt.month - 1
    if m == 0:
        y -= 1
        m = 12
    return y, m


async def monthly_auto_report_loop() -> None:
    while True:
        try:
            init_extra_db()
            now = datetime.now(timezone.utc)
            year, month = previous_month_year_month(now)
            if now.day == 1:
                conn = get_db()
                exists = conn.execute(
                    "SELECT 1 FROM monthly_report_log WHERE year = ? AND month = ?",
                    (year, month),
                ).fetchone()
                conn.close()
                if not exists:
                    file_path, caption = monthly_report(month, year)
                    for admin_id in ADMIN_IDS:
                        try:
                            await bot.send_document(
                                admin_id,
                                BufferedInputFile(file_path.read_bytes(), filename=file_path.name),
                                caption=caption,
                            )
                        except Exception:
                            logger.exception("Failed to auto send monthly report to %s", admin_id)
                    conn = get_db()
                    conn.execute(
                        "INSERT OR REPLACE INTO monthly_report_log (year, month, sent_at) VALUES (?, ?, ?)",
                        (year, month, utc_now_iso()),
                    )
                    conn.commit()
                    conn.close()
        except Exception:
            logger.exception("monthly_auto_report_loop failed")
        await asyncio.sleep(21600)


# ============================================================
# FAVORITES API
# ============================================================

@web_router.get("/api/shop/favorites")
async def api_favorites(request: web.Request) -> web.Response:
    user_id = safe_int(request.query.get("user_id"))
    conn = get_db()
    rows = conn.execute(
        "SELECT product_id FROM favorites WHERE user_id = ? ORDER BY id DESC",
        (user_id,),
    ).fetchall()
    conn.close()
    return web.json_response({"favorites": [safe_int(r["product_id"]) for r in rows]})


@web_router.post("/api/shop/favorites/toggle")
async def api_toggle_favorite(request: web.Request) -> web.Response:
    try:
        data = await request.json()
    except Exception:
        data = {}
    user_id = safe_int(data.get("user_id"))
    product_id = safe_int(data.get("product_id"))
    if user_id <= 0 or product_id <= 0:
        return web.json_response({"ok": False, "error": "bad_params"}, status=400)

    conn = get_db()
    row = conn.execute(
        "SELECT id FROM favorites WHERE user_id = ? AND product_id = ?",
        (user_id, product_id),
    ).fetchone()
    if row:
        conn.execute("DELETE FROM favorites WHERE id = ?", (row["id"],))
        favorite = False
    else:
        conn.execute(
            "INSERT OR IGNORE INTO favorites (user_id, product_id, created_at) VALUES (?, ?, ?)",
            (user_id, product_id, utc_now_iso()),
        )
        favorite = True
    conn.commit()
    conn.close()
    return web.json_response({"ok": True, "favorite": favorite})


# ============================================================
# BETTER WEBAPP HTML
# ============================================================

def build_shop_html() -> str:
    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{html.escape(SHOP_BRAND)}</title>
<script src="https://telegram.org/js/telegram-web-app.js"></script>
<style>
:root {{
  --bg1:#fffaf3;
  --bg2:#fff4f8;
  --bg3:#fdf7ef;
  --ivory:#fffaf6;
  --cream:#f7efe3;
  --champ:#ead9ba;
  --gold:#c9a45f;
  --gold2:#9e7731;
  --text:#241913;
  --muted:#7b695f;
  --pink:#b10f63;
  --pink2:#7f184f;
  --violet:#6d3db4;
  --violet2:#4b267f;
  --glass:rgba(255,255,255,.68);
  --stroke:rgba(205,170,107,.23);
  --dark:#17110d;
  --ok:#26654c;
  --softShadow:0 22px 54px rgba(95,74,43,.08);
}}
* {{ box-sizing:border-box; }}
html,body {{ margin:0; padding:0; min-height:100%; }}
body {{
  font-family:Inter,Arial,sans-serif;
  color:var(--text);
  background:
    radial-gradient(circle at top left, rgba(214,185,118,.14), transparent 26%),
    radial-gradient(circle at top right, rgba(133,57,141,.14), transparent 24%),
    linear-gradient(135deg, var(--bg1), var(--bg2), var(--bg3));
  overflow-x:hidden;
}}
.petal-layer {{
  position:fixed;
  inset:0;
  pointer-events:none;
  z-index:0;
  overflow:hidden;
}}
.petal {{
  position:absolute;
  top:-12vh;
  border-radius:55% 45% 60% 40% / 50% 45% 55% 50%;
  opacity:.92;
  filter:drop-shadow(0 10px 14px rgba(95,41,102,.18));
  animation-name:fallPetal, swayPetal;
  animation-timing-function:linear, ease-in-out;
  animation-iteration-count:infinite, infinite;
}}
@keyframes fallPetal {{
  0% {{ transform:translateY(-12vh) rotate(0deg); }}
  100% {{ transform:translateY(120vh) rotate(360deg); }}
}}
@keyframes swayPetal {{
  0%,100% {{ margin-left:-8px; }}
  50% {{ margin-left:18px; }}
}}
.shell {{
  position:relative;
  z-index:1;
  max-width:1240px;
  margin:0 auto;
  padding:18px 14px 38px;
}}
.hero {{
  position:relative;
  overflow:hidden;
  padding:26px 22px 24px;
  border-radius:0 0 32px 32px;
  background:
    linear-gradient(135deg, rgba(255,255,255,.76), rgba(255,250,242,.55)),
    radial-gradient(circle at 20% 15%, rgba(199,164,99,.16), transparent 26%),
    radial-gradient(circle at 82% 18%, rgba(177,38,99,.16), transparent 24%),
    radial-gradient(circle at 62% 40%, rgba(109,61,180,.10), transparent 26%);
  box-shadow:var(--softShadow);
  border:1px solid rgba(255,255,255,.66);
  backdrop-filter:blur(16px);
}}
.hero::after {{
  content:'';
  position:absolute;
  inset:auto -5% -40px auto;
  width:320px;
  height:140px;
  background:radial-gradient(circle, rgba(201,164,95,.16), transparent 66%);
}}
.brand {{
  font-size:40px;
  line-height:1;
  font-weight:900;
  letter-spacing:.18em;
  text-transform:uppercase;
  background:linear-gradient(180deg, #f3dfb1, #c9a45f, #9e7731);
  -webkit-background-clip:text;
  -webkit-text-fill-color:transparent;
}}
.hero-sub {{
  margin-top:8px;
  color:var(--muted);
  font-size:14px;
  letter-spacing:.03em;
}}
.hero-row {{
  display:flex;
  gap:12px;
  flex-wrap:wrap;
  margin-top:16px;
}}
.hero-chip {{
  padding:10px 14px;
  border-radius:999px;
  background:rgba(255,255,255,.66);
  border:1px solid rgba(201,164,95,.20);
  color:#5f4e45;
  font-weight:700;
  box-shadow:0 8px 22px rgba(88,69,37,.04);
}}
.layout {{
  display:grid;
  grid-template-columns:minmax(0,1.56fr) minmax(320px,.92fr);
  gap:16px;
  margin-top:16px;
}}
.panel {{
  background:var(--glass);
  backdrop-filter:blur(14px);
  border:1px solid rgba(255,255,255,.72);
  border-radius:26px;
  box-shadow:var(--softShadow);
  overflow:hidden;
}}
.panel-head {{
  padding:18px 18px 8px;
  font-size:24px;
  font-weight:900;
}}
.panel-sub {{
  padding:0 18px 14px;
  color:var(--muted);
  font-size:13px;
}}
.toolbar {{
  padding:0 16px 14px;
  display:grid;
  grid-template-columns:1.2fr .8fr .75fr .75fr;
  gap:10px;
}}
.search,.select,.num {{
  width:100%;
  border:none;
  background:rgba(255,255,255,.84);
  border:1px solid var(--stroke);
  border-radius:16px;
  padding:12px 14px;
  color:var(--text);
  outline:none;
}}
.chips {{
  display:flex;
  gap:8px;
  flex-wrap:wrap;
  padding:0 16px 14px;
}}
.chip {{
  border:none;
  cursor:pointer;
  padding:10px 14px;
  border-radius:999px;
  background:rgba(255,255,255,.86);
  color:#5b493c;
  font-weight:800;
  border:1px solid var(--stroke);
}}
.chip.active {{
  color:#fff;
  background:linear-gradient(180deg, var(--gold), var(--gold2));
}}
.grid {{
  display:grid;
  grid-template-columns:repeat(2, minmax(0,1fr));
  gap:14px;
  padding:0 14px 16px;
}}
.card {{
  position:relative;
  background:rgba(255,255,255,.80);
  border-radius:22px;
  border:1px solid rgba(255,255,255,.86);
  box-shadow:0 14px 28px rgba(89,67,34,.06);
  overflow:hidden;
}}
.photo-wrap {{
  position:relative;
  padding:12px 12px 0;
}}
.photo {{
  position:relative;
  aspect-ratio:1/1.08;
  border-radius:18px;
  overflow:hidden;
  background:
    radial-gradient(circle at 30% 20%, rgba(201,164,95,.16), transparent 28%),
    linear-gradient(135deg, #fff, #f7eee3, #fff6f9);
}}
.photo img {{
  width:100%;
  height:100%;
  object-fit:cover;
  display:block;
}}
.no-photo {{
  width:100%;
  height:100%;
  display:flex;
  align-items:center;
  justify-content:center;
  color:#8f7a69;
  font-weight:700;
  letter-spacing:.05em;
}}
.badges {{
  position:absolute;
  top:18px;
  left:18px;
  display:flex;
  gap:7px;
  flex-wrap:wrap;
  z-index:2;
}}
.badge {{
  padding:6px 10px;
  border-radius:999px;
  font-size:10px;
  font-weight:900;
  color:#fff;
  letter-spacing:.05em;
  text-transform:uppercase;
  box-shadow:0 8px 16px rgba(59,33,16,.16);
}}
.badge-new {{ background:linear-gradient(180deg,#6d3db4,#4b267f); }}
.badge-hit {{ background:linear-gradient(180deg,#b10f63,#7f184f); }}
.badge-sale {{ background:linear-gradient(180deg,#d7ac58,#9e7731); }}
.badge-limited {{ background:linear-gradient(180deg,#2f2119,#17110d); }}
.badge-low_stock {{ background:linear-gradient(180deg,#9f326e,#6d2350); }}
.fav {{
  position:absolute;
  top:18px;
  right:18px;
  width:42px;
  height:42px;
  border:none;
  border-radius:50%;
  cursor:pointer;
  background:rgba(255,255,255,.88);
  box-shadow:0 10px 20px rgba(78,52,35,.10);
  font-size:18px;
}}
.fav.active {{
  color:#b10f63;
}}
.card-inner {{
  padding:14px 14px 16px;
}}
.title {{
  font-size:17px;
  font-weight:900;
  line-height:1.35;
}}
.desc {{
  margin-top:8px;
  min-height:40px;
  color:var(--muted);
  font-size:13px;
  line-height:1.45;
}}
.price-row {{
  display:flex;
  align-items:flex-end;
  gap:10px;
  flex-wrap:wrap;
  margin-top:12px;
}}
.price {{
  font-size:22px;
  font-weight:900;
}}
.old {{
  color:#a18f84;
  text-decoration:line-through;
  font-size:13px;
}}
.stock {{
  margin-top:10px;
  font-size:12px;
  color:#5f5249;
  font-weight:700;
}}
.stock.low {{
  color:#8c255a;
}}
.stock.out {{
  color:#8c1c1c;
}}
.sizes {{
  display:flex;
  gap:7px;
  flex-wrap:wrap;
  margin-top:12px;
}}
.size-btn {{
  border:none;
  cursor:pointer;
  padding:8px 12px;
  border-radius:999px;
  background:#fff;
  border:1px solid var(--stroke);
  font-weight:800;
  color:#5a493d;
}}
.size-btn.active {{
  color:#fff;
  background:linear-gradient(180deg, var(--gold), var(--gold2));
}}
.qty-row {{
  margin-top:12px;
  display:flex;
  align-items:center;
  gap:8px;
}}
.qty-ctl {{
  display:flex;
  align-items:center;
  gap:6px;
  background:#fff;
  border:1px solid var(--stroke);
  border-radius:14px;
  padding:6px;
}}
.qty-ctl button {{
  border:none;
  background:#f8f1e7;
  width:32px;
  height:32px;
  border-radius:10px;
  cursor:pointer;
  font-size:18px;
  font-weight:800;
}}
.qty-ctl span {{
  min-width:22px;
  text-align:center;
  font-weight:900;
}}
.actions {{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:10px;
  margin-top:14px;
}}
.btn-black,.btn-gold,.btn-white {{
  border:none;
  cursor:pointer;
  border-radius:16px;
  padding:13px 14px;
  font-weight:900;
  transition:transform .12s ease, opacity .12s ease;
}}
.btn-black:active,.btn-gold:active,.btn-white:active {{
  transform:translateY(1px);
}}
.btn-black {{
  color:#fff;
  background:var(--dark);
}}
.btn-gold {{
  color:#fff;
  background:linear-gradient(180deg, var(--gold), var(--gold2));
}}
.btn-white {{
  color:#59493d;
  background:#fff;
  border:1px solid var(--stroke);
}}
.btn-black[disabled], .btn-gold[disabled], .btn-white[disabled] {{
  cursor:not-allowed;
  opacity:.52;
}}
.side-stack {{
  display:flex;
  flex-direction:column;
  gap:16px;
}}
.cart-list,.review-list {{
  display:flex;
  flex-direction:column;
  gap:10px;
  padding:0 16px 16px;
}}
.cart-item,.review-item {{
  background:rgba(255,255,255,.86);
  border:1px solid rgba(205,170,107,.18);
  border-radius:18px;
  padding:12px;
}}
.summary {{
  padding:0 16px 16px;
}}
.summary-row {{
  display:flex;
  justify-content:space-between;
  gap:10px;
  margin-bottom:8px;
}}
.mini-muted {{
  color:var(--muted);
  font-size:12px;
}}
.notice {{
  position:fixed;
  left:50%;
  bottom:18px;
  transform:translateX(-50%);
  background:#1d140f;
  color:#fff;
  padding:12px 16px;
  border-radius:999px;
  z-index:10;
  opacity:0;
  transition:opacity .18s ease;
  box-shadow:0 12px 26px rgba(38,25,18,.22);
}}
.notice.show {{
  opacity:1;
}}
.empty {{
  padding:14px;
  color:#7f6f63;
}}
@media(max-width:980px) {{
  .layout {{ grid-template-columns:1fr; }}
}}
@media(max-width:760px) {{
  .toolbar {{ grid-template-columns:1fr 1fr; }}
  .grid {{ grid-template-columns:1fr; }}
  .brand {{ font-size:32px; }}
}}
@media(max-width:520px) {{
  .toolbar {{ grid-template-columns:1fr; }}
  .actions {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class="petal-layer" id="petalLayer"></div>

<div class="shell">
  <section class="hero">
    <div class="brand">{html.escape(SHOP_BRAND)}</div>
    <div class="hero-sub">Premium collection inside Telegram</div>
    <div class="hero-row">
      <div class="hero-chip">Luxury spring feeling</div>
      <div class="hero-chip">Soft gold / glassmorphism</div>
      <div class="hero-chip">Delivery & checkout in Telegram</div>
    </div>
  </section>

  <section class="layout">
    <div class="panel">
      <div class="panel-head">Каталог</div>
      <div class="panel-sub">Поиск, категории, размеры, цена, избранное, статусы товара.</div>

      <div class="toolbar">
        <input id="searchInput" class="search" placeholder="Поиск по названию">
        <select id="sizeFilter" class="select"><option value="">Все размеры</option></select>
        <input id="minPrice" class="num" type="number" min="0" placeholder="Цена от">
        <input id="maxPrice" class="num" type="number" min="0" placeholder="Цена до">
      </div>

      <div class="chips" id="categoryChips"></div>
      <div class="grid" id="productGrid"></div>
    </div>

    <div class="side-stack">
      <div class="panel">
        <div class="panel-head">Корзина</div>
        <div class="panel-sub">Оформление продолжается прямо в боте.</div>
        <div class="cart-list" id="cartList"></div>
        <div class="summary">
          <div class="summary-row"><span>Всего</span><b id="sumQty">0</b></div>
          <div class="summary-row"><span>Сумма</span><b id="sumAmount">0 сум</b></div>
          <button class="btn-black" style="width:100%;margin-top:8px" id="checkoutBtn">Оформить заказ</button>
          <button class="btn-white" style="width:100%;margin-top:8px" id="clearBtn">Очистить корзину</button>
        </div>
      </div>

      <div class="panel">
        <div class="panel-head">Избранное ❤️</div>
        <div class="panel-sub">Быстрый доступ к понравившимся товарам.</div>
        <div class="cart-list" id="favoriteList"></div>
      </div>

      <div class="panel">
        <div class="panel-head">Отзывы</div>
        <div class="panel-sub">Показываются опубликованные отзывы.</div>
        <div class="review-list" id="reviewList"></div>
      </div>
    </div>
  </section>
</div>

<div class="notice" id="notice"></div>

<script>
const tg = window.Telegram?.WebApp || null;
if (tg) {{
  tg.ready();
  tg.expand();
}}

const userId = tg?.initDataUnsafe?.user?.id || 0;
const lang = (new URLSearchParams(window.location.search).get('lang') || 'ru').toLowerCase();

const TEXT = {{
  ru: {{
    all: 'Все',
    new: 'Новинки',
    hits: 'Хиты',
    sale: 'Скидки',
    limited: 'Limited',
    school: 'Школа',
    casual: 'Casual',
    add: 'В корзину',
    buy: 'Купить сейчас',
    out: 'Нет в наличии',
    low: 'Осталось',
    emptyCart: 'Корзина пуста',
    emptyFav: 'Пока нет избранного',
    emptyReviews: 'Пока нет отзывов',
    searchEmpty: 'Ничего не найдено',
    left: 'шт',
    stock: 'В наличии',
    badge_new: 'Новинка',
    badge_hit: 'Хит',
    badge_sale: 'Скидка',
    badge_limited: 'Limited',
    badge_low_stock: 'Скоро закончится',
    added: 'Добавлено в корзину',
    removed: 'Удалено',
    cleared: 'Корзина очищена',
    favAdded: 'Добавлено в избранное',
    favRemoved: 'Убрано из избранного',
    openInTelegram: 'Открой через Telegram',
  }},
  uz: {{
    all: 'Barchasi',
    new: 'Yangi',
    hits: 'Hitlar',
    sale: 'Chegirma',
    limited: 'Limited',
    school: 'Maktab',
    casual: 'Casual',
    add: "Savatchaga",
    buy: "Hozir sotib olish",
    out: "Tugagan",
    low: "Qoldi",
    emptyCart: "Savatcha bo'sh",
    emptyFav: "Hozircha yo'q",
    emptyReviews: "Hozircha sharh yo'q",
    searchEmpty: "Hech narsa topilmadi",
    left: 'ta',
    stock: 'Mavjud',
    badge_new: 'Yangi',
    badge_hit: 'Hit',
    badge_sale: 'Chegirma',
    badge_limited: 'Limited',
    badge_low_stock: 'Tugab bormoqda',
    added: "Savatchaga qo'shildi",
    removed: "O'chirildi",
    cleared: "Savatcha tozalandi",
    favAdded: "Saralanganlarga qo'shildi",
    favRemoved: "Saralanganlardan olib tashlandi",
    openInTelegram: "Telegram orqali oching",
  }}
}};
const L = TEXT[lang] || TEXT.ru;

const state = {{
  products: [],
  cart: [],
  reviews: [],
  favoriteIds: [],
  category: 'all',
  size: '',
  minPrice: '',
  maxPrice: '',
  search: '',
  qtyById: {{}},
  sizeById: {{}},
}};

const categories = [
  ['all', L.all],
  ['new', L.new],
  ['hits', L.hits],
  ['sale', L.sale],
  ['limited', L.limited],
  ['school', L.school],
  ['casual', L.casual]
];

function esc(v) {{
  return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}}[s]));
}}
function fmtSum(v) {{
  return Number(v || 0).toLocaleString('ru-RU') + ' сум';
}}
function showNotice(text) {{
  const el = document.getElementById('notice');
  el.textContent = text;
  el.classList.add('show');
  setTimeout(() => el.classList.remove('show'), 1800);
}}
function sendData(payload) {{
  if (!tg) {{
    showNotice(L.openInTelegram);
    return;
  }}
  tg.sendData(JSON.stringify(payload));
}}

function createPetals() {{
  const layer = document.getElementById('petalLayer');
  layer.innerHTML = '';
  const colors = [
    'linear-gradient(180deg,#7a43c6,#4b267f)',
    'linear-gradient(180deg,#b10f63,#7f184f)',
    'linear-gradient(180deg,#8e4ed1,#6d3db4)',
    'linear-gradient(180deg,#c63b79,#92225a)',
    'linear-gradient(180deg,#d88ad7,#8a4eb7)'
  ];
  for (let i = 0; i < 28; i++) {{
    const p = document.createElement('div');
    const size = 10 + Math.random() * 18;
    p.className = 'petal';
    p.style.left = (Math.random() * 100) + 'vw';
    p.style.width = size + 'px';
    p.style.height = (size * 0.74) + 'px';
    p.style.background = colors[Math.floor(Math.random() * colors.length)];
    const d1 = 8 + Math.random() * 12;
    const d2 = 2.4 + Math.random() * 2.8;
    const delay = Math.random() * 8;
    p.style.animationDuration = d1 + 's,' + d2 + 's';
    p.style.animationDelay = delay + 's,' + delay + 's';
    p.style.opacity = String(0.58 + Math.random() * 0.35);
    layer.appendChild(p);
  }}
}}

async function loadProducts() {{
  const res = await fetch(`/api/shop/products?lang=${{encodeURIComponent(lang)}}`);
  state.products = await res.json();
  state.products.forEach(p => {{
    if (!state.qtyById[p.id]) state.qtyById[p.id] = 1;
    if (!state.sizeById[p.id]) state.sizeById[p.id] = (Array.isArray(p.sizes_list) && p.sizes_list[0]) ? p.sizes_list[0] : '';
  }});
  buildSizeFilter();
  renderCategoryChips();
  renderProducts();
}}

async function loadCart() {{
  if (!userId) return;
  const res = await fetch(`/api/shop/cart?user_id=${{encodeURIComponent(userId)}}&lang=${{encodeURIComponent(lang)}}`);
  const data = await res.json();
  state.cart = data.items || [];
  renderCart();
}}

async function loadReviews() {{
  const res = await fetch('/api/shop/reviews');
  state.reviews = await res.json();
  renderReviews();
}}

async function loadFavorites() {{
  if (!userId) return;
  const res = await fetch(`/api/shop/favorites?user_id=${{encodeURIComponent(userId)}}`);
  const data = await res.json();
  state.favoriteIds = Array.isArray(data.favorites) ? data.favorites : [];
  renderFavoriteList();
  renderProducts();
}}

async function toggleFavorite(productId) {{
  if (!userId) {{
    showNotice(L.openInTelegram);
    return;
  }}
  const res = await fetch('/api/shop/favorites/toggle', {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify({{ user_id: userId, product_id: productId }})
  }});
  const data = await res.json();
  if (data.favorite) showNotice(L.favAdded); else showNotice(L.favRemoved);
  await loadFavorites();
}}

function buildSizeFilter() {{
  const select = document.getElementById('sizeFilter');
  const sizes = new Set();
  state.products.forEach(p => (p.sizes_list || []).forEach(s => sizes.add(String(s))));
  select.innerHTML = `<option value="">${{lang === 'uz' ? "Barcha o'lchamlar" : "Все размеры"}}</option>` +
    [...sizes].sort().map(s => `<option value="${{esc(s)}}">${{esc(s)}}</option>`).join('');
  select.value = state.size || '';
}}

function renderCategoryChips() {{
  const box = document.getElementById('categoryChips');
  box.innerHTML = '';
  categories.forEach(([key, title]) => {{
    const btn = document.createElement('button');
    btn.className = 'chip' + (state.category === key ? ' active' : '');
    btn.textContent = title;
    btn.onclick = () => {{
      state.category = key;
      renderCategoryChips();
      renderProducts();
    }};
    box.appendChild(btn);
  }});
}}

function badgeClass(key) {{
  return 'badge badge-' + key;
}}
function badgeText(key) {{
  return L['badge_' + key] || key;
}}
function isFavorite(productId) {{
  return state.favoriteIds.includes(productId);
}}
function filteredProducts() {{
  const q = String(state.search || '').trim().toLowerCase();
  const minPrice = state.minPrice === '' ? null : Number(state.minPrice);
  const maxPrice = state.maxPrice === '' ? null : Number(state.maxPrice);

  return state.products.filter(p => {{
    const title = String(p.title || '').toLowerCase();
    const bySearch = !q || title.includes(q);
    const byCategory =
      state.category === 'all' ||
      p.category_slug === state.category ||
      (state.category === 'new' && (p.badges || []).includes('new')) ||
      (state.category === 'hits' && (p.badges || []).includes('hit')) ||
      (state.category === 'sale' && (p.badges || []).includes('sale')) ||
      (state.category === 'limited' && (p.badges || []).includes('limited'));

    const sizes = Array.isArray(p.sizes_list) ? p.sizes_list.map(x => String(x)) : [];
    const bySize = !state.size || sizes.includes(String(state.size));
    const byMin = minPrice === null || Number(p.price || 0) >= minPrice;
    const byMax = maxPrice === null || Number(p.price || 0) <= maxPrice;
    return bySearch && byCategory && bySize && byMin && byMax;
  }});
}}

function stockLabel(p) {{
  const stock = Number(p.stock_qty || 0);
  if (stock <= 0) return `<div class="stock out">${{L.out}}</div>`;
  if (stock <= {LOW_STOCK_THRESHOLD}) return `<div class="stock low">${{L.low}} ${{stock}} ${{L.left}}</div>`;
  return `<div class="stock">${{L.stock}}: ${{stock}}</div>`;
}}

function renderProducts() {{
  const grid = document.getElementById('productGrid');
  const items = filteredProducts();
  if (!items.length) {{
    grid.innerHTML = `<div class="card"><div class="card-inner">${{L.searchEmpty}}</div></div>`;
    return;
  }}
  grid.innerHTML = '';
  items.forEach(p => {{
    const card = document.createElement('div');
    card.className = 'card';

    const selectedSize = state.sizeById[p.id] || ((p.sizes_list || [])[0] || '');
    const qty = Number(state.qtyById[p.id] || 1);
    const canBuy = !!p.can_buy;

    card.innerHTML = `
      <div class="photo-wrap">
        <div class="badges">
          ${(p.badges || []).map(k => `<span class="${{badgeClass(k)}}">${{esc(badgeText(k))}}</span>`).join('')}
        </div>
        <button class="fav ${{isFavorite(p.id) ? 'active' : ''}}" title="favorite">❤</button>
        <div class="photo">
          ${
            p.photo_url
              ? `<img src="${{esc(p.photo_url)}}" alt="${{esc(p.title)}}">`
              : `<div class="no-photo">${{esc('{SHOP_BRAND}')}}</div>`
          }
        </div>
      </div>
      <div class="card-inner">
        <div class="title">${{esc(p.title)}}</div>
        <div class="desc">${{esc(p.description || '')}}</div>
        <div class="price-row">
          <div class="price">${{fmtSum(p.price)}}</div>
          ${Number(p.old_price || 0) > 0 ? `<div class="old">${{fmtSum(p.old_price)}}</div>` : ``}
        </div>
        ${{stockLabel(p)}}
        <div class="sizes">
          ${(p.sizes_list || []).map(s => `<button class="size-btn ${{String(s)===String(selectedSize)?'active':''}}" data-size="${{esc(s)}}">${{esc(s)}}</button>`).join('')}
        </div>
        <div class="qty-row">
          <div class="qty-ctl">
            <button class="minus">−</button>
            <span class="qty-val">${{qty}}</span>
            <button class="plus">+</button>
          </div>
          <div class="mini-muted">ID: #${{p.id}}</div>
        </div>
        <div class="actions">
          <button class="btn-black addBtn" ${{canBuy ? '' : 'disabled'}}>${{canBuy ? L.add : L.out}}</button>
          <button class="btn-gold buyBtn" ${{canBuy ? '' : 'disabled'}}>${{canBuy ? L.buy : L.out}}</button>
        </div>
      </div>
    `;

    const favBtn = card.querySelector('.fav');
    favBtn.onclick = () => toggleFavorite(p.id);

    card.querySelectorAll('.size-btn').forEach(btn => {{
      btn.onclick = () => {{
        state.sizeById[p.id] = btn.dataset.size || '';
        renderProducts();
      }};
    }});

    const qtyVal = card.querySelector('.qty-val');
    card.querySelector('.minus').onclick = () => {{
      state.qtyById[p.id] = Math.max(1, Number(state.qtyById[p.id] || 1) - 1);
      qtyVal.textContent = String(state.qtyById[p.id]);
    }};
    card.querySelector('.plus').onclick = () => {{
      const current = Number(state.qtyById[p.id] || 1);
      const maxByStock = Math.max(1, Number(p.stock_qty || 1));
      state.qtyById[p.id] = Math.min(maxByStock, current + 1);
      qtyVal.textContent = String(state.qtyById[p.id]);
    }};

    card.querySelector('.addBtn').onclick = () => {{
      sendData({{
        action: 'add_to_cart',
        product_id: p.id,
        qty: Number(state.qtyById[p.id] || 1),
        size: state.sizeById[p.id] || ''
      }});
      showNotice(L.added);
      setTimeout(loadCart, 450);
    }};
    card.querySelector('.buyBtn').onclick = () => {{
      sendData({{
        action: 'buy_now',
        product_id: p.id,
        qty: Number(state.qtyById[p.id] || 1),
        size: state.sizeById[p.id] || ''
      }});
      showNotice(L.buy);
      setTimeout(loadCart, 450);
    }};

    grid.appendChild(card);
  }});
}}

function renderCart() {{
  const box = document.getElementById('cartList');
  const qtyEl = document.getElementById('sumQty');
  const amountEl = document.getElementById('sumAmount');
  box.innerHTML = '';
  let totalQty = 0;
  let totalAmount = 0;

  (state.cart || []).forEach(item => {{
    totalQty += Number(item.qty || 0);
    totalAmount += Number(item.subtotal || 0);
    const div = document.createElement('div');
    div.className = 'cart-item';
    div.innerHTML = `
      <div><b>${{esc(item.product_name)}}</b></div>
      <div class="mini-muted" style="margin-top:4px">
        ${{item.size ? esc(item.size) + ' | ' : ''}}${{item.qty}} × ${{fmtSum(item.price)}}
      </div>
      <div style="display:flex;justify-content:space-between;align-items:center;margin-top:8px;gap:10px">
        <b>${{fmtSum(item.subtotal)}}</b>
        <button class="btn-white removeBtn" style="padding:10px 12px;border-radius:12px">Удалить</button>
      </div>
    `;
    div.querySelector('.removeBtn').onclick = () => {{
      sendData({{ action: 'remove_from_cart', cart_id: item.cart_id }});
      showNotice(L.removed);
      setTimeout(loadCart, 450);
    }};
    box.appendChild(div);
  }});

  if (!(state.cart || []).length) {{
    box.innerHTML = `<div class="cart-item">${{L.emptyCart}}</div>`;
  }}
  qtyEl.textContent = String(totalQty);
  amountEl.textContent = fmtSum(totalAmount);
}}

function renderFavoriteList() {{
  const box = document.getElementById('favoriteList');
  const favProducts = state.products.filter(p => state.favoriteIds.includes(p.id));
  if (!favProducts.length) {{
    box.innerHTML = `<div class="cart-item">${{L.emptyFav}}</div>`;
    return;
  }}
  box.innerHTML = '';
  favProducts.slice(0, 10).forEach(p => {{
    const div = document.createElement('div');
    div.className = 'cart-item';
    div.innerHTML = `
      <div style="display:flex;justify-content:space-between;gap:10px">
        <div>
          <b>${{esc(p.title)}}</b>
          <div class="mini-muted" style="margin-top:4px">${{fmtSum(p.price)}}</div>
        </div>
        <button class="btn-white openBtn" style="padding:10px 12px;border-radius:12px">Открыть</button>
      </div>
    `;
    div.querySelector('.openBtn').onclick = () => {{
      const el = [...document.querySelectorAll('.title')].find(x => x.textContent.trim() === p.title.trim());
      if (el) el.scrollIntoView({{ behavior: 'smooth', block: 'center' }});
    }};
    box.appendChild(div);
  }});
}}

function renderReviews() {{
  const box = document.getElementById('reviewList');
  if (!state.reviews.length) {{
    box.innerHTML = `<div class="review-item">${{L.emptyReviews}}</div>`;
    return;
  }}
  box.innerHTML = '';
  state.reviews.slice(0, 12).forEach(r => {{
    const div = document.createElement('div');
    div.className = 'review-item';
    div.innerHTML = `
      <div>${{'⭐'.repeat(Math.max(1, Math.min(5, Number(r.rating || 5))))}}</div>
      <div style="font-weight:900;margin-top:6px">${{esc(r.customer_name || 'Client')}}</div>
      <div class="mini-muted" style="margin-top:6px;line-height:1.45">${{esc(r.text || '')}}</div>
    `;
    box.appendChild(div);
  }});
}}

document.getElementById('searchInput').addEventListener('input', e => {{
  state.search = e.target.value || '';
  renderProducts();
}});
document.getElementById('sizeFilter').addEventListener('change', e => {{
  state.size = e.target.value || '';
  renderProducts();
}});
document.getElementById('minPrice').addEventListener('input', e => {{
  state.minPrice = e.target.value || '';
  renderProducts();
}});
document.getElementById('maxPrice').addEventListener('input', e => {{
  state.maxPrice = e.target.value || '';
  renderProducts();
}});
document.getElementById('checkoutBtn').onclick = () => sendData({{ action: 'checkout' }});
document.getElementById('clearBtn').onclick = () => {{
  sendData({{ action: 'clear_cart' }});
  showNotice(L.cleared);
  setTimeout(loadCart, 450);
}};

createPetals();
loadProducts().then(() => loadFavorites());
loadCart();
loadReviews();
</script>
</body>
</html>
"""


# ============================================================
# IMPROVED ADMIN SHELL (REDEFINITION)
# ============================================================

def admin_template(title: str, body: str) -> str:
    token = html.escape(ADMIN_PANEL_TOKEN)
    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{html.escape(title)}</title>
<style>
:root {{
  --gold:#c9a45f;
  --gold2:#9e7731;
  --bg:#fff9f3;
  --bg2:#fff4f8;
  --text:#231812;
  --muted:#78695d;
  --stroke:#eadbc4;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0;
  font-family:Inter,Arial,sans-serif;
  color:var(--text);
  background:linear-gradient(135deg,var(--bg),var(--bg2),#fdf7ef);
}}
.wrap {{ max-width:1280px; margin:0 auto; padding:20px 14px 34px; }}
.top {{
  background:rgba(255,255,255,.78);
  border:1px solid rgba(255,255,255,.80);
  backdrop-filter:blur(12px);
  border-radius:26px;
  padding:20px;
  box-shadow:0 16px 34px rgba(89,67,34,.06);
}}
.brand {{
  font-size:34px;
  font-weight:900;
  letter-spacing:.16em;
  text-transform:uppercase;
  background:linear-gradient(180deg,#f4e0ae,var(--gold),var(--gold2));
  -webkit-background-clip:text;
  -webkit-text-fill-color:transparent;
}}
.sub {{ margin-top:6px; color:var(--muted); }}
.nav {{
  display:flex;
  flex-wrap:wrap;
  gap:10px;
  margin-top:14px;
}}
.nav a {{
  text-decoration:none;
  color:#fff;
  background:linear-gradient(180deg,var(--gold),var(--gold2));
  padding:10px 14px;
  border-radius:999px;
  font-weight:800;
}}
.card {{
  margin-top:16px;
  background:rgba(255,255,255,.84);
  border:1px solid rgba(255,255,255,.84);
  border-radius:24px;
  padding:18px;
  box-shadow:0 12px 24px rgba(89,67,34,.05);
}}
.stats {{
  display:grid;
  grid-template-columns:repeat(4,minmax(0,1fr));
  gap:12px;
}}
.stat {{
  background:#fff;
  border:1px solid var(--stroke);
  border-radius:18px;
  padding:16px;
}}
.label {{ color:var(--muted); font-size:12px; }}
.value {{ font-size:28px; font-weight:900; margin-top:8px; }}
table {{
  width:100%;
  border-collapse:collapse;
}}
th,td {{
  padding:10px 8px;
  border-bottom:1px solid var(--stroke);
  text-align:left;
  vertical-align:top;
}}
th {{
  background:#faf4ea;
}}
input, textarea, select {{
  width:100%;
  border:1px solid var(--stroke);
  border-radius:14px;
  padding:11px 12px;
  background:#fff;
  color:var(--text);
  outline:none;
}}
textarea {{ min-height:120px; resize:vertical; }}
.grid-2 {{
  display:grid;
  grid-template-columns:repeat(2,minmax(0,1fr));
  gap:12px;
}}
.grid-3 {{
  display:grid;
  grid-template-columns:repeat(3,minmax(0,1fr));
  gap:12px;
}}
.btn {{
  display:inline-block;
  text-decoration:none;
  border:none;
  cursor:pointer;
  background:linear-gradient(180deg,var(--gold),var(--gold2));
  color:#fff;
  padding:11px 16px;
  border-radius:14px;
  font-weight:900;
}}
.btn.white {{
  color:#57483d;
  background:#fff;
  border:1px solid var(--stroke);
}}
.tag {{
  display:inline-flex;
  align-items:center;
  gap:6px;
  padding:7px 10px;
  border-radius:999px;
  background:#fff;
  border:1px solid var(--stroke);
  font-size:12px;
  font-weight:800;
  color:#5a4a3e;
}}
.inline-form {{
  display:flex;
  gap:8px;
  align-items:center;
  flex-wrap:wrap;
}}
.success {{
  padding:10px 12px;
  border-radius:14px;
  background:#edf8ef;
  color:#21563c;
  margin-bottom:14px;
  border:1px solid #c8e7d0;
}}
@media(max-width:980px) {{
  .stats {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
  .grid-3 {{ grid-template-columns:1fr; }}
}}
@media(max-width:740px) {{
  .grid-2 {{ grid-template-columns:1fr; }}
  table {{ display:block; overflow:auto; white-space:nowrap; }}
}}
</style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div class="brand">{html.escape(SHOP_BRAND)} Admin</div>
    <div class="sub">Web control panel</div>
    <div class="nav">
      <a href="/admin?token={token}">Dashboard</a>
      <a href="/admin/orders?token={token}">Orders</a>
      <a href="/admin/products?token={token}">Products</a>
      <a href="/admin/reviews?token={token}">Reviews</a>
      <a href="/shop?lang=ru" target="_blank">Open Shop</a>
    </div>
  </div>

  {body}
</div>

<script>
(function() {{
  const token = {json.dumps(ADMIN_PANEL_TOKEN)};

  if (location.pathname === '/admin/products') {{
    const headerRow = document.querySelector('table thead tr');
    if (headerRow && !headerRow.querySelector('.edit-head')) {{
      const th = document.createElement('th');
      th.className = 'edit-head';
      th.textContent = 'Edit';
      headerRow.appendChild(th);
    }}
    document.querySelectorAll('table tbody tr').forEach(tr => {{
      if (tr.querySelector('.edit-cell')) return;
      const firstCell = tr.children[0];
      if (!firstCell) return;
      const id = (firstCell.textContent || '').replace(/[^0-9]/g, '');
      if (!id) return;
      const td = document.createElement('td');
      td.className = 'edit-cell';
      td.innerHTML = `<a class="btn white" href="/admin/products/${{id}}/edit?token=${{encodeURIComponent(token)}}">Edit</a>`;
      tr.appendChild(td);
    }});
  }}

  if (location.pathname === '/admin/reviews') {{
    document.querySelectorAll('[data-review-id]').forEach(box => {{
      const id = box.getAttribute('data-review-id');
      const published = box.getAttribute('data-review-published') === '1';
      if (published) return;
      const footer = document.createElement('div');
      footer.style.marginTop = '10px';
      footer.innerHTML = `<form method="post" action="/admin/reviews/${{id}}/publish?token=${{encodeURIComponent(token)}}" style="display:inline-block"><button class="btn" type="submit">Publish</button></form>`;
      box.appendChild(footer);
    }});
  }}

  if (location.pathname === '/admin/orders') {{
    const headerRow = document.querySelector('table thead tr');
    if (headerRow && !headerRow.querySelector('.actions-head')) {{
      const th1 = document.createElement('th');
      th1.textContent = 'Статус';
      headerRow.appendChild(th1);
      const th2 = document.createElement('th');
      th2.textContent = 'Оплата';
      headerRow.appendChild(th2);
    }}

    document.querySelectorAll('table tbody tr').forEach(tr => {{
      if (tr.querySelector('.js-order-actions')) return;
      const idCell = tr.children[0];
      if (!idCell) return;
      const id = (idCell.textContent || '').replace(/[^0-9]/g, '');
      if (!id) return;

      const statusTd = document.createElement('td');
      statusTd.className = 'js-order-actions';
      statusTd.innerHTML = `
        <form method="post" action="/admin/orders/${{id}}/status?token=${{encodeURIComponent(token)}}" class="inline-form">
          <select name="status">
            <option value="new">new</option>
            <option value="confirmed">confirmed</option>
            <option value="paid">paid</option>
            <option value="sent">sent</option>
            <option value="delivered">delivered</option>
            <option value="cancelled">cancelled</option>
          </select>
          <button class="btn white" type="submit">Save</button>
        </form>
      `;

      const payTd = document.createElement('td');
      payTd.innerHTML = `
        <form method="post" action="/admin/orders/${{id}}/payment?token=${{encodeURIComponent(token)}}" class="inline-form">
          <select name="payment_status">
            <option value="pending">pending</option>
            <option value="paid">paid</option>
            <option value="failed">failed</option>
            <option value="cancelled">cancelled</option>
            <option value="refunded">refunded</option>
          </select>
          <button class="btn white" type="submit">Save</button>
        </form>
      `;
      tr.appendChild(statusTd);
      tr.appendChild(payTd);
    }});
  }}
}})();
</script>
</body>
</html>
"""


# ============================================================
# REAL PRODUCT EDIT PAGE + SAVE
# ============================================================

def parse_checkbox(form: web.Request, key: str) -> int:
    return 1 if form.get(key) in {"1", "on", "true", "yes"} else 0


def product_edit_form_html(row: sqlite3.Row, token: str, success: str = "") -> str:
    checked_published = "checked" if safe_int(row["is_published"]) else ""
    checked_new = "checked" if safe_int(row["is_new"]) else ""
    checked_hit = "checked" if safe_int(row["is_hit"]) else ""
    checked_limited = "checked" if safe_int(row["is_limited"]) else ""
    success_html = f"<div class='success'>{html.escape(success)}</div>" if success else ""
    return f"""
<div class="card">
  <h2 style="margin:0 0 16px">Edit product #{row['id']}</h2>
  {success_html}
  <form method="post" action="/admin/products/{row['id']}/edit?token={token}">
    <div class="grid-2">
      <div>
        <label>title_ru</label>
        <input name="title_ru" value="{html.escape(row['title_ru'] or '')}" required>
      </div>
      <div>
        <label>title_uz</label>
        <input name="title_uz" value="{html.escape(row['title_uz'] or '')}" required>
      </div>
    </div>

    <div class="grid-2" style="margin-top:12px">
      <div>
        <label>description_ru</label>
        <textarea name="description_ru">{html.escape(row['description_ru'] or '')}</textarea>
      </div>
      <div>
        <label>description_uz</label>
        <textarea name="description_uz">{html.escape(row['description_uz'] or '')}</textarea>
      </div>
    </div>

    <div class="grid-3" style="margin-top:12px">
      <div>
        <label>price</label>
        <input type="number" min="0" name="price" value="{safe_int(row['price'])}" required>
      </div>
      <div>
        <label>old_price</label>
        <input type="number" min="0" name="old_price" value="{safe_int(row['old_price'])}">
      </div>
      <div>
        <label>discount_percent</label>
        <input type="number" min="0" max="100" name="discount_percent" value="{safe_int(row['discount_percent'])}">
      </div>
    </div>

    <div class="grid-3" style="margin-top:12px">
      <div>
        <label>sizes (через запятую)</label>
        <input name="sizes" value="{html.escape(', '.join(parse_json_list(row['sizes_json'])))}">
      </div>
      <div>
        <label>stock_qty</label>
        <input type="number" min="0" name="stock_qty" value="{safe_int(row['stock_qty'])}">
      </div>
      <div>
        <label>category_slug</label>
        <select name="category_slug">
          {''.join(f'<option value="{html.escape(cat)}" {"selected" if row["category_slug"] == cat else ""}>{html.escape(cat)}</option>' for cat in CATEGORY_SLUGS)}
        </select>
      </div>
    </div>

    <div class="grid-3" style="margin-top:12px">
      <div>
        <label>photo_file_id</label>
        <input name="photo_file_id" value="{html.escape(row['photo_file_id'] or '')}">
      </div>
      <div>
        <label>sort_order</label>
        <input type="number" name="sort_order" value="{safe_int(row['sort_order'])}">
      </div>
      <div>
        <label>is_published</label>
        <div class="tag"><input type="checkbox" name="is_published" value="1" {checked_published}> Published</div>
      </div>
    </div>

    <div class="grid-3" style="margin-top:12px">
      <div class="tag"><input type="checkbox" name="is_new" value="1" {checked_new}> is_new</div>
      <div class="tag"><input type="checkbox" name="is_hit" value="1" {checked_hit}> is_hit</div>
      <div class="tag"><input type="checkbox" name="is_limited" value="1" {checked_limited}> is_limited</div>
    </div>

    <div style="margin-top:16px;display:flex;gap:10px;flex-wrap:wrap">
      <button class="btn" type="submit">Save product</button>
      <a class="btn white" href="/admin/products?token={token}">Back to list</a>
    </div>
  </form>
</div>
"""


@web_router.get("/admin/products/{product_id}/edit")
async def admin_edit_product_page(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)

    product_id = safe_int(request.match_info.get("product_id"))
    row = get_product(product_id)
    if not row:
        return web.Response(text="Product not found", status=404)

    body = product_edit_form_html(row, html.escape(request.query.get("token", "")))
    return web.Response(text=admin_template(f"Edit Product #{product_id}", body), content_type="text/html")


@web_router.post("/admin/products/{product_id}/edit")
async def admin_edit_product_save(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)

    product_id = safe_int(request.match_info.get("product_id"))
    row = get_product(product_id)
    if not row:
        return web.Response(text="Product not found", status=404)

    form = await request.post()
    title_ru = (form.get("title_ru") or "").strip()
    title_uz = (form.get("title_uz") or "").strip()
    description_ru = (form.get("description_ru") or "").strip()
    description_uz = (form.get("description_uz") or "").strip()
    price = max(0, safe_int(form.get("price")))
    old_price = max(0, safe_int(form.get("old_price")))
    discount_percent = max(0, min(100, safe_int(form.get("discount_percent"))))
    stock_qty = max(0, safe_int(form.get("stock_qty")))
    category_slug = (form.get("category_slug") or "casual").strip()
    if category_slug not in CATEGORY_SLUGS:
        category_slug = "casual"
    photo_file_id = (form.get("photo_file_id") or "").strip()
    sort_order = safe_int(form.get("sort_order"), 100)

    sizes_raw = (form.get("sizes") or "").strip()
    sizes_json = json.dumps([s.strip() for s in sizes_raw.split(",") if s.strip()], ensure_ascii=False)

    conn = get_db()
    conn.execute(
        """
        UPDATE products
        SET title_ru = ?, title_uz = ?, description_ru = ?, description_uz = ?,
            price = ?, old_price = ?, sizes_json = ?, stock_qty = ?, category_slug = ?,
            is_published = ?, is_new = ?, is_hit = ?, is_limited = ?, discount_percent = ?,
            photo_file_id = ?, sort_order = ?, updated_at = ?
        WHERE id = ?
        """,
        (
            title_ru,
            title_uz,
            description_ru,
            description_uz,
            price,
            old_price,
            sizes_json,
            stock_qty,
            category_slug,
            parse_checkbox(form, "is_published"),
            parse_checkbox(form, "is_new"),
            parse_checkbox(form, "is_hit"),
            parse_checkbox(form, "is_limited"),
            discount_percent,
            photo_file_id,
            sort_order,
            utc_now_iso(),
            product_id,
        ),
    )
    conn.commit()
    conn.close()

    updated = get_product(product_id)
    body = product_edit_form_html(updated, html.escape(request.query.get("token", "")), success="Product saved successfully.")
    return web.Response(text=admin_template(f"Edit Product #{product_id}", body), content_type="text/html")


# ============================================================
# WEB ADMIN ACTIONS FOR ORDERS / REVIEWS
# ============================================================

@web_router.post("/admin/reviews/{review_id}/publish")
async def admin_publish_review_page(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)
    review_id = safe_int(request.match_info.get("review_id"))
    conn = get_db()
    conn.execute(
        "UPDATE reviews SET is_published = 1, updated_at = ? WHERE id = ?",
        (utc_now_iso(), review_id),
    )
    conn.commit()
    conn.close()
    raise web.HTTPFound(location=f"/admin/reviews?token={request.query.get('token','')}")


@web_router.post("/admin/orders/{order_id}/status")
async def admin_order_status_page(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)

    order_id = safe_int(request.match_info.get("order_id"))
    form = await request.post()
    status = (form.get("status") or "").strip()
    if status not in ORDER_STATUSES:
        return web.Response(text="Bad status", status=400)

    conn = get_db()
    conn.execute(
        "UPDATE orders SET status = ?, updated_at = ? WHERE id = ?",
        (status, utc_now_iso(), order_id),
    )
    conn.commit()
    conn.close()
    raise web.HTTPFound(location=f"/admin/orders?token={request.query.get('token','')}")


@web_router.post("/admin/orders/{order_id}/payment")
async def admin_order_payment_page(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)

    order_id = safe_int(request.match_info.get("order_id"))
    form = await request.post()
    payment_status = (form.get("payment_status") or "").strip()
    if payment_status not in PAYMENT_STATUSES:
        return web.Response(text="Bad payment status", status=400)

    conn = get_db()
    conn.execute(
        "UPDATE orders SET payment_status = ?, updated_at = ? WHERE id = ?",
        (payment_status, utc_now_iso(), order_id),
    )
    conn.commit()
    conn.close()
    raise web.HTTPFound(location=f"/admin/orders?token={request.query.get('token','')}")


# ============================================================
# REDEFINE REVIEWS PAGE BODY FORMAT TO SUPPORT PUBLISH JS
# ============================================================

@web_router.get("/admin/reviews_v2")
async def admin_reviews_page_v2(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)
    conn = get_db()
    rows = conn.execute("SELECT * FROM reviews ORDER BY id DESC LIMIT 200").fetchall()
    conn.close()
    body = "<div class='card'>" + "".join(
        f"""
        <div data-review-id="{row['id']}" data-review-published="{safe_int(row['is_published'])}" style="padding:12px 0;border-bottom:1px solid #eee4d2">
          <b>#{row['id']}</b> | {html.escape(row['customer_name'] or 'Client')} |
          {'⭐' * max(1, min(5, safe_int(row['rating'], 5)))}<br>
          <div style="margin-top:8px;line-height:1.5">{html.escape(row['text'] or '')}</div>
          <small>{'Опубликован' if safe_int(row['is_published']) else 'На модерации'}</small>
        </div>
        """
        for row in rows
    ) + "</div>"
    return web.Response(text=admin_template("Reviews", body), content_type="text/html")


# ============================================================
# OPTIONAL OVERRIDE FOR CONTACT / ADMIN TITLES IN WEB PAGES
# ============================================================

@web_router.get("/admin/products_v2")
async def admin_products_page_v2(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)
    rows = get_products(published_only=False)
    tr = []
    for row in rows:
        badges = ", ".join(stock_badges(row)) or "—"
        tr.append(
            f"<tr>"
            f"<td>#{row['id']}</td>"
            f"<td><b>{html.escape(row['title_ru'])}</b><br><span style='color:#7a6d61'>{html.escape(row['title_uz'])}</span></td>"
            f"<td>{html.escape(row['category_slug'])}</td>"
            f"<td>{fmt_sum(row['price'])}</td>"
            f"<td>{row['stock_qty']}</td>"
            f"<td>{'Да' if safe_int(row['is_published']) else 'Нет'}</td>"
            f"<td>{html.escape(badges)}</td>"
            f"<td><a class='btn white' href='/admin/products/{row['id']}/edit?token={html.escape(request.query.get('token',''))}'>Edit</a></td>"
            f"</tr>"
        )
    body = (
        "<div class='card'>"
        "<table><thead><tr>"
        "<th>ID</th><th>Товар</th><th>Категория</th><th>Цена</th><th>Остаток</th><th>Опубликован</th><th>Бейджи</th><th>Edit</th>"
        "</tr></thead><tbody>"
        + "".join(tr) +
        "</tbody></table></div>"
    )
    return web.Response(text=admin_template("Products", body), content_type="text/html")


# ============================================================
# OPTIONAL: ENHANCED CHECKOUT SUMMARY HELPER
# ============================================================

def build_checkout_items_preview(user_id: int) -> str:
    rows = get_cart_rows(user_id)
    if not rows:
        return "—"
    lines = []
    for idx, row in enumerate(rows, start=1):
        size = f" | {html.escape(row['size'])}" if row["size"] else ""
        qty = safe_int(row["qty"])
        price = safe_int(row["price"])
        lines.append(f"{idx}. <b>{html.escape(row['title_ru'])}</b>{size} — {qty} × {fmt_sum(price)}")
    return "\n".join(lines)


# ============================================================
# CORRECTED CHECKOUT COMMENT HANDLER
# ЕСЛИ СТАВИШЬ ЭТУ ЧАСТЬ В ФАЙЛ, УДАЛИ СТАРЫЙ checkout_comment_handler
# ============================================================

@checkout_router.message(CheckoutState.comment)
async def checkout_comment_handler_v2(message: Message, state: FSMContext) -> None:
    if await maybe_cancel(message, state):
        return
    comment = (message.text or "").strip()
    if comment == t(message.from_user.id, "skip"):
        comment = ""
    await state.update_data(comment=comment)
    data = await state.get_data()
    total_qty, total_amount = cart_totals(message.from_user.id)

    text = (
        f"{t(message.from_user.id, 'checkout_summary')}\n\n"
        f"<b>Имя:</b> {html.escape(data.get('customer_name', '—'))}\n"
        f"<b>Телефон:</b> {html.escape(data.get('customer_phone', '—'))}\n"
        f"<b>Город:</b> {html.escape(data.get('city', '—'))}\n"
        f"<b>Способ доставки:</b> {delivery_label(message.from_user.id, data.get('delivery_method', 'pickup'))}\n"
        f"<b>Адрес:</b> {html.escape(data.get('delivery_address', '—'))}\n"
        f"<b>Способ оплаты:</b> {html.escape(data.get('payment_method', 'cash'))}\n"
        f"<b>Комментарий:</b> {html.escape(comment or '—')}\n\n"
        f"<b>Товары:</b>\n{build_checkout_items_preview(message.from_user.id)}\n\n"
        f"<b>Количество:</b> {total_qty}\n"
        f"<b>Сумма:</b> {fmt_sum(total_amount)}\n\n"
        f"{t(message.from_user.id, 'checkout_confirm_hint')}"
    )
    await state.set_state(CheckoutState.confirm)
    await message.answer(text, reply_markup=yes_no_keyboard(message.from_user.id))


# ============================================================
# CORRECTED ADMIN COMMANDS WITH IMMEDIATE NOTIFY
# ЕСЛИ СТАВИШЬ ЭТУ ЧАСТЬ В ФАЙЛ, УДАЛИ СТАРЫЕ /set_order_status И /set_payment_status
# ============================================================

@admin_router.message(Command("set_order_status"))
async def admin_set_order_status_v2(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return
    parts = (message.text or "").split()
    if len(parts) != 3:
        await message.answer("Используй: /set_order_status 15 sent")
        return
    order_id = safe_int(parts[1])
    status = parts[2].strip()
    if status not in ORDER_STATUSES:
        await message.answer("Допустимые статусы: new, confirmed, paid, sent, delivered, cancelled")
        return
    conn = get_db()
    conn.execute("UPDATE orders SET status = ?, updated_at = ? WHERE id = ?", (status, utc_now_iso(), order_id))
    conn.commit()
    row = conn.execute("SELECT user_id FROM orders WHERE id = ?", (order_id,)).fetchone()
    conn.close()

    if row:
        user_id = safe_int(row["user_id"])
        lang = get_user_lang(user_id)
        text = order_client_text(lang, order_id, status)
        if text:
            try:
                await bot.send_message(user_id, text)
            except Exception:
                logger.exception("Failed to notify user about order status")
    await message.answer(f"Статус заказа #{order_id} обновлён: {status}")


@admin_router.message(Command("set_payment_status"))
async def admin_set_payment_status_v2(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return
    parts = (message.text or "").split()
    if len(parts) != 3:
        await message.answer("Используй: /set_payment_status 15 paid")
        return
    order_id = safe_int(parts[1])
    payment_status = parts[2].strip()
    if payment_status not in PAYMENT_STATUSES:
        await message.answer("Допустимые статусы: pending, paid, failed, cancelled, refunded")
        return
    conn = get_db()
    conn.execute("UPDATE orders SET payment_status = ?, updated_at = ? WHERE id = ?", (payment_status, utc_now_iso(), order_id))
    conn.commit()
    row = conn.execute("SELECT user_id FROM orders WHERE id = ?", (order_id,)).fetchone()
    conn.close()

    if row:
        user_id = safe_int(row["user_id"])
        lang = get_user_lang(user_id)
        text = payment_client_text(lang, order_id, payment_status)
        if text:
            try:
                await bot.send_message(user_id, text)
            except Exception:
                logger.exception("Failed to notify user about payment status")
    await message.answer(f"Статус оплаты заказа #{order_id} обновлён: {payment_status}")


# ============================================================
# CORRECTED /publish_review COMMAND
# ЕСЛИ СТАВИШЬ ЭТУ ЧАСТЬ В ФАЙЛ, УДАЛИ СТАРЫЙ /publish_review
# ============================================================

@admin_router.message(Command("publish_review"))
async def admin_publish_review_v2(message: Message) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(t(message.from_user.id, "not_admin"))
        return
    parts = (message.text or "").split()
    if len(parts) != 2:
        await message.answer("Используй: /publish_review 5")
        return
    review_id = safe_int(parts[1])
    conn = get_db()
    conn.execute("UPDATE reviews SET is_published = 1, updated_at = ? WHERE id = ?", (utc_now_iso(), review_id))
    conn.commit()
    conn.close()
    await message.answer(f"Отзыв #{review_id} опубликован.")


# ============================================================
# BETTER PRODUCTS LIST PAGE (UNIQUE ROUTE)
# ============================================================

@web_router.get("/admin/products-full")
async def admin_products_full_page(request: web.Request) -> web.Response:
    if not admin_allowed(request):
        return web.Response(text="Access denied", status=403)

    rows = get_products(published_only=False)
    token = html.escape(request.query.get("token", ""))
    body = "<div class='card'><table><thead><tr><th>ID</th><th>RU / UZ</th><th>Категория</th><th>Цена</th><th>Old</th><th>Sizes</th><th>Stock</th><th>Published</th><th>Flags</th><th>Edit</th></tr></thead><tbody>"
    for row in rows:
        flags = []
        if safe_int(row["is_new"]): flags.append("new")
        if safe_int(row["is_hit"]): flags.append("hit")
        if safe_int(row["is_limited"]): flags.append("limited")
        if safe_int(row["discount_percent"]): flags.append(f"-{safe_int(row['discount_percent'])}%")
        body += (
            f"<tr>"
            f"<td>#{row['id']}</td>"
            f"<td><b>{html.escape(row['title_ru'])}</b><br><span style='color:#7a6d61'>{html.escape(row['title_uz'])}</span></td>"
            f"<td>{html.escape(row['category_slug'])}</td>"
            f"<td>{fmt_sum(row['price'])}</td>"
            f"<td>{fmt_sum(row['old_price']) if safe_int(row['old_price']) else '—'}</td>"
            f"<td>{html.escape(', '.join(parse_json_list(row['sizes_json'])) or '—')}</td>"
            f"<td>{row['stock_qty']}</td>"
            f"<td>{'Да' if safe_int(row['is_published']) else 'Нет'}</td>"
            f"<td>{html.escape(', '.join(flags) or '—')}</td>"
            f"<td><a class='btn white' href='/admin/products/{row['id']}/edit?token={token}'>Edit</a></td>"
            f"</tr>"
        )
    body += "</tbody></table></div>"
    return web.Response(text=admin_template("Products Full", body), content_type="text/html")


# ============================================================
# REPLACE FINAL STARTUP BLOCK WITH THIS ONE
# ============================================================

def create_web_app() -> web.Application:
    app = web.Application(client_max_size=10 * 1024 * 1024)
    app.add_routes(web_router)
    return app


def register_routers() -> None:
    dp.include_router(user_router)
    dp.include_router(cart_router)
    dp.include_router(checkout_router)
    dp.include_router(orders_router)
    dp.include_router(reviews_router)
    dp.include_router(admin_router)
    dp.include_router(fallback_router)


async def stock_watch_loop() -> None:
    while True:
        try:
            await notify_admins_low_stock()
        except Exception:
            logger.exception("stock_watch_loop failed")
        await asyncio.sleep(3600)


async def main() -> None:
    init_db()
    init_extra_db()
    seed_products_if_empty()
    register_routers()

    app = create_web_app()
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    logger.info("Web server started on port %s", PORT)

    asyncio.create_task(stock_watch_loop())
    asyncio.create_task(order_notify_loop())
    asyncio.create_task(monthly_auto_report_loop())

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
